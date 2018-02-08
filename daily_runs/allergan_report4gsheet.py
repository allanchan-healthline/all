from delivery_helpers import *
from gsheet_gdrive_api import *

import openpyxl as opx
from datetime import datetime, timedelta

############################################
folder_id = '0B71ox_2Qc7gmSS1fbUJKTXEyams'
############################################

############################################
# Get data
############################################

yesterday_date = datetime.now().date() - timedelta(days=1)
month_start_date = yesterday_date.replace(day=1)
report_start_date = str(month_start_date.year) + str(month_start_date.month).zfill(2) + str(month_start_date.day).zfill(2)
report_end_date = str(yesterday_date.year) + str(yesterday_date.month).zfill(2) + str(yesterday_date.day).zfill(2)

(o1_original, o2_labeled, o3_summary, o4_only_billable) = allergan_report(report_start_date, report_end_date)

############################################
# Create a new google sheet
############################################

gdrive_service = get_gdrive_service()
gsheet_service = get_gsheet_service()

file_data = {'name': 'allergan_report_' + report_start_date + '_' + report_end_date,
             'parents': [folder_id],
             'mimeType': 'application/vnd.google-apps.spreadsheet'}
file_info = gdrive_service.files().create(body=file_data).execute()
file_id = file_info['id']

############################################
# Upload 3 sheets
# Reference: https://developers.google.com/sheets/api/samples/sheet#add_a_sheet
############################################

upload_dict = {'original': o1_original, 'labeled': o2_labeled, 'summary': o3_summary}

for sheet_name in ['summary', 'labeled', 'original']:
    df = upload_dict[sheet_name]
    df = df.fillna('')
    values = [df.columns.tolist()] + df.values.tolist()

    # Add formulas
    if sheet_name == 'summary':
        header = values[0]

        i_device = header.index('Device')
        i_billable = header.index('Billable Imps')
        col_billable = opx.utils.get_column_letter(i_billable + 1)
        i_dfp = header.index('DFP Imps')
        col_dfp = opx.utils.get_column_letter(i_dfp + 1)
        i_billability = header.index('Billability')
        col_billability = opx.utils.get_column_letter(i_billability + 1)
        i_discrepancy = header.index('Discrepancy')

        for i in range(len(values)):
            if values[i][i_device] == 'Total':
                str_row = str(i + 1)
                values[i][i_billability] = '=' + col_billable + str_row + '/' + col_dfp + str_row
                values[i][i_discrepancy] = '=1-' + col_billability + str_row
    
    request_body = {'requests': [{'addSheet': {'properties': {'title': sheet_name,
                                                              'gridProperties': {'rowCount': len(values),
                                                                                 'columnCount': len(values[0])}}}}]}
    result = gsheet_service.spreadsheets().batchUpdate(spreadsheetId=file_id, body=request_body).execute()
    sheet_id = result['replies'][0]['addSheet']['properties']['sheetId']
    
    sheet_range = sheet_name + '!A1:' + opx.utils.get_column_letter(len(values[0])) + str(len(values))
    result = gsheet_service.spreadsheets().values().update(spreadsheetId=file_id, range=sheet_range,
                                                           valueInputOption='USER_ENTERED', body={'values': values}).execute()

    if sheet_name == 'summary':
        summary_sheet_id = sheet_id
        summary_values = values

############################################
# Delete 'Sheet 1'
# Reference: https://developers.google.com/sheets/api/samples/sheet#delete_a_sheet
############################################

ss_metadata = gsheet_service.spreadsheets().get(spreadsheetId=file_id).execute()
for sheet_metadata in ss_metadata['sheets']:
    if sheet_metadata['properties']['title'] == 'Sheet1':
        sheet_id = sheet_metadata['properties']['sheetId']
        break

request_body = {'requests': [{'deleteSheet': {'sheetId': sheet_id}}]}
result = gsheet_service.spreadsheets().batchUpdate(spreadsheetId=file_id, body=request_body).execute()

############################################
# Format the summary sheet
############################################

n_row = len(summary_values)
header = summary_values[0]
n_col = len(header)

############################################
value_formatting = []

def make_vf_dict(i_row_start, n_row, i_col_start, n_col, type, pattern):
    return {'repeatCell': {'range': {'sheetId': summary_sheet_id,
                                     'startRowIndex': i_row_start,
                                     'endRowIndex': i_row_start + n_row,
                                     'startColumnIndex': i_col_start,
                                     'endColumnIndex': i_col_start + n_col},
                           'cell': {'userEnteredFormat': {'numberFormat': {'type': type,
                                                                           'pattern': pattern}}},
                           'fields': 'userEnteredFormat.numberFormat'}}

value_formatting.append(make_vf_dict(1, n_row - 1, header.index('Valid and Viewable Impressions'), 2, 'NUMBER', '###,###,##0'))
value_formatting.append(make_vf_dict(1, n_row - 1, header.index('Viewability'), 2, 'NUMBER', '#0.0%'))
value_formatting.append(make_vf_dict(1, n_row - 1, header.index('Impressions Analyzed'), 3, 'NUMBER', '###,###,##0'))
value_formatting.append(make_vf_dict(1, n_row - 1, header.index('Billability'), 2, 'NUMBER', '#0.0%'))
value_formatting.append(make_vf_dict(1, n_row - 1, header.index('GroupM Payable Impressions'), 1, 'NUMBER', '###,###,##0'))

############################################
default = {'red': 1, 'green': 1, 'blue': 1, 'alpha': 1}
grey = {'red': .949, 'green': .949, 'blue': .949, 'alpha': 1}
green = {'red': .851, 'green': .918, 'blue': .827, 'alpha': 1}
blue = {'red': .788, 'green': .855, 'blue': .973, 'alpha': 1}
sky = {'red': .855, 'green': .933, 'blue': .953, 'alpha': 1}
orange = {'red': .988, 'green': .898, 'blue': .804, 'alpha': 1}

i_device = header.index('Device')
i_size = header.index('Size')

colors = []

for i in range(len(summary_values)):
    if i == 0:
        colors += [{'values': [{'userEnteredFormat': {'backgroundColor': grey}}] * n_col}]
    elif summary_values[i][i_size] == 'Total':
        if summary_values[i][i_device] == 'Desktop & Tablet':
            colors += [{'values': [{'userEnteredFormat': {'backgroundColor': green}}] * n_col}]
        elif summary_values[i][i_device] == 'Mobile':
            colors += [{'values': [{'userEnteredFormat': {'backgroundColor': blue}}] * n_col}]
        else:
            colors += [{'values': [{'userEnteredFormat': {'backgroundColor': grey}}] * n_col}]
    elif summary_values[i][i_device] == 'Total':
        colors += [{'values': [{'userEnteredFormat': {'backgroundColor': orange}}] * n_col}]
    else:
        colors += [{'values': [{'userEnteredFormat': {'backgroundColor': default}}] * n_col}]

color_formatting = [{'updateCells': {'rows': colors,
                                     'fields': 'userEnteredFormat.backgroundColor',
                                     'range': {'sheetId': summary_sheet_id,
                                               'startRowIndex': 0,
                                               'endRowIndex': n_row,
                                               'startColumnIndex': 0,
                                               'endColumnIndex': n_col}}}]

############################################
wrap = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'wrapStrategy': 'WRAP'}}] * len(header)}],
                         'fields': 'userEnteredFormat.wrapStrategy',
                         'range': {'sheetId': summary_sheet_id,
                                   'startRowIndex': 0,
                                   'endRowIndex': 1,
                                   'startColumnIndex': 0,
                                   'endColumnIndex': n_col}}}]

############################################
freeze = [{'updateSheetProperties': {'properties': {'sheetId': summary_sheet_id,
                                                    'gridProperties': {'frozenRowCount': 1}},
                                     'fields': 'gridProperties(frozenRowCount)'}}]

############################################
width = [{'updateDimensionProperties': {'range': {'sheetId': summary_sheet_id,
                                                  'dimension': 'COLUMNS',
                                                  'startIndex': 0,
                                                  'endIndex': len(header)},
                                        'properties': {'pixelSize': 77},
                                        'fields': 'pixelSize'}}]

############################################
all_formatting_requests = value_formatting + color_formatting + wrap + freeze + width
result = gsheet_service.spreadsheets().batchUpdate(spreadsheetId=file_id,
                                                   body={'requests': all_formatting_requests}).execute()


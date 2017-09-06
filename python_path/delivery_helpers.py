from path2pickles import *
from NEW_helpers import *
from gsheet_gdrive_api import *

import os
import shutil
import time
import pickle

import requests

import dfareporting_utils
from oauth2client import client
import googleapiclient

from googleads import dfp
from googleads import errors

import pandas as pd
import csv
import re
import openpyxl as opx
from datetime import datetime, date

import pytz
from pytz import timezone

pd.options.mode.chained_assignment = None  # default='warn'

#####################################################
# DFP delivery via API
#####################################################

def run_dfp_mtd_all_query(last_delivery_date):

    end_date = last_delivery_date
    start_date = last_delivery_date.replace(day=1)
    output_file_name = 'temp_dfp_mtd_all.csv'

    # Report query id
    saved_query_id = '10000261152'

    # Initialize a client
    client = dfp.DfpClient.LoadFromStorage()

    # Initialize appropriate service.
    report_service = client.GetService('ReportService', version='v201705')

    # Initialize a DataDownloader.
    report_downloader = client.GetDataDownloader(version='v201705')

    # Create statement object to filter for an order.
    values = [{'key': 'id',
               'value': {'xsi_type': 'NumberValue',
                         'value': saved_query_id}}]
    query = 'WHERE id = :id'
    statement = dfp.FilterStatement(query, values, 1)

    response = report_service.getSavedQueriesByStatement(statement.ToStatement())

    if 'results' in response:
        saved_query = response['results'][0]

        if saved_query['isCompatibleWithApiVersion']:
            report_job = {}

            # Set report query and optionally modify it.
            report_job['reportQuery'] = saved_query['reportQuery']

            report_job['reportQuery']['startDate']['year'] = start_date.year
            report_job['reportQuery']['startDate']['month'] = start_date.month
            report_job['reportQuery']['startDate']['day'] = start_date.day

            report_job['reportQuery']['endDate']['year'] = end_date.year
            report_job['reportQuery']['endDate']['month'] = end_date.month
            report_job['reportQuery']['endDate']['day'] = end_date.day

            try:
                # Run the report and wait for it to finish.
                report_job_id = report_downloader.WaitForReport(report_job)
            except errors.DfpReportError as e:
                print('Failed to generate report. Error was: %s' % e)

            # Download report data.
            with open(output_file_name, 'wb') as report_file:
                report_downloader.DownloadReportToFile(report_job_id, 'CSV_DUMP', report_file, use_gzip_compression=False)

        else:
            print('The query specified is not compatible with the API version.')

    # Clean up
    cols = ['Dimension.DATE',
            'Dimension.ADVERTISER_NAME',
            'Dimension.ORDER_NAME',
            'Dimension.LINE_ITEM_NAME',
            'Dimension.CREATIVE_NAME',
            'Dimension.CREATIVE_SIZE',
            'Dimension.AD_UNIT_NAME',
            'DimensionAttribute.LINE_ITEM_COST_PER_UNIT',
            'CF[6995]_Value',
            'CF[7115]_Value',
            'Column.TOTAL_LINE_ITEM_LEVEL_IMPRESSIONS',
            'Column.TOTAL_LINE_ITEM_LEVEL_CLICKS']

    col_rename_dict = {'Dimension.DATE': 'Date',
                       'Dimension.ADVERTISER_NAME': 'Advertiser',
                       'Dimension.ORDER_NAME': 'Order',
                       'Dimension.LINE_ITEM_NAME': 'Line item',
                       'Dimension.CREATIVE_NAME': 'Creative',
                       'Dimension.CREATIVE_SIZE': 'Creative size',
                       'Dimension.AD_UNIT_NAME': 'Ad unit',
                       'DimensionAttribute.LINE_ITEM_COST_PER_UNIT': 'Rate ($)',
                       'CF[6995]_Value': 'DAS Line Item Name',
                       'CF[7115]_Value': '3rd Party Creative ID',
                       'Column.TOTAL_LINE_ITEM_LEVEL_IMPRESSIONS': 'Total impressions',
                       'Column.TOTAL_LINE_ITEM_LEVEL_CLICKS': 'Total clicks'}

    dfp_mtd_all = pd.read_csv(output_file_name , encoding='utf-8')
    dfp_mtd_all['DimensionAttribute.LINE_ITEM_COST_PER_UNIT'] = dfp_mtd_all['DimensionAttribute.LINE_ITEM_COST_PER_UNIT'] / 1000000.0
    dfp_mtd_all = dfp_mtd_all[cols].rename(columns=col_rename_dict)
    dfp_mtd_all['Date'] = [datetime.strptime(d, '%Y-%m-%d').date() for d in dfp_mtd_all['Date']]

    os.remove(output_file_name)

    return dfp_mtd_all

def get_dfp_today_delivery():

    today_date = datetime.now(tz=pytz.utc).astimezone(timezone('US/Eastern')).date()

    output_file_name = 'temp_dfp_today_delivery.csv'

    ########################################################
    # Get Order Name, Line Item Name, Creative Name, DAS Line Item Name
    ########################################################

    dfp_client = dfp.DfpClient.LoadFromStorage()
    filter_statement = {'query': "WHERE ORDER_NAME LIKE '%BBR%'"}

    report_job = {
        'reportQuery': {
            'dimensions': ['ORDER_NAME',
                           'LINE_ITEM_NAME',
                           'CREATIVE_NAME'],
            'statement': filter_statement,
            'columns': ['AD_SERVER_IMPRESSIONS'],
            'dateRangeType': 'CUSTOM_DATE',
            'startDate': {'year': today_date.year,
                          'month': today_date.month,
                          'day': today_date.day},
            'endDate': {'year': today_date.year,
                        'month': today_date.month,
                        'day': today_date.day},
            'customFieldIds': [6995]
        }
    }

    report_downloader = dfp_client.GetDataDownloader(version='v201702')
    try:
        generated_at = datetime.now(tz=pytz.utc).astimezone(timezone('US/Eastern'))
        report_job_id = report_downloader.WaitForReport(report_job)
    except errors.DfpReportError as e:
        print('Failed to generate report. Error was: %s' % e)

    with open(output_file_name, 'wb') as report_file:
        report_downloader.DownloadReportToFile(report_job_id, 'CSV_DUMP', report_file, use_gzip_compression=False)

    ########################################################
    # Clean up
    ########################################################

    cols = ['Dimension.ORDER_NAME',
            'Dimension.LINE_ITEM_NAME',
            'Dimension.CREATIVE_NAME',
            'CF[6995]_Value',
            'Column.AD_SERVER_IMPRESSIONS']

    col_rename_dict = {'Dimension.ORDER_NAME': 'Order',
                       'Dimension.LINE_ITEM_NAME': 'Line item',
                       'Dimension.CREATIVE_NAME': 'Creative',
                       'CF[6995]_Value': 'DAS Line Item Name',
                       'Column.AD_SERVER_IMPRESSIONS': 'Ad Server Impressions'}

    df = pd.read_csv(output_file_name, encoding='utf-8')
    df = df[cols].rename(columns=col_rename_dict)

    os.remove(output_file_name)
    return (generated_at, df)

#####################################################
# DFP MTD All query either with a pre-generated file
# or via API
#####################################################

def get_dfp_mtd_all(last_delivery_date, path_csv, emailed_csv):

    if path_csv is None:
        dfp_mtd_all = run_dfp_mtd_all_query(last_delivery_date)
    else:
        if emailed_csv:
            skiprows = 7
        else:
            skiprows = 9
        dfp_mtd_all = pd.read_csv(path_csv, skiprows=skiprows, skipfooter=1, engine='python', encoding='utf-8')
    return dfp_mtd_all

#####################################################
# UVs
#####################################################

def get_microsite_uvs(site, mo_year, path_excel, cpuv_goals_sheet, mnt_uv_tracker_tabs):
    wb = opx.load_workbook(path_excel, data_only=True)
    cpuv_goals = get_cpuv_goals(mo_year[1], cpuv_goals_sheet)

    ###########################################################
    # Make a list of microsite tab names
    ###########################################################

    if site == 'HL':
        temp_sheets = cpuv_goals['HL Report Tab Name'].drop_duplicates().values.tolist()
        sheets = []
        for temp_sheet in temp_sheets:
            if (temp_sheet is not None) and (temp_sheet != ''):
                sheets += temp_sheet.split(', ')
    elif site == 'MNT':
        sheets = mnt_uv_tracker_tabs
    elif site == 'Drugs.com':
        sheets = []
        for sheet in wb.get_sheet_names():
            if '(TS)' in sheet:
                continue
            if '(Treatment Seekers)' in sheet:
                continue
            if '(Treatment Seeker)' in sheet:
                continue
            sheets.append(sheet)
    else:
        sheets = wb.get_sheet_names()

    ###########################################################
    # Specify date col and uv col
    ###########################################################

    date_col_num = 2
    if site in ['HL', 'MNT']:
        uv_col_num = 4
    else:
        uv_col_num = 3

    ###########################################################
    # Collect UVs
    ###########################################################

    uv_list = []  # This will be a list of [site, tab name, date, uvs]
    if site == 'MNT':
        site = 'Medical News Today'

    for sheet in sheets:
        next_sheet = False
        
        try:
            ws = wb.get_sheet_by_name(sheet)
        except KeyError:
            print('FIX! A tab named ' + sheet + ' is not in the ' + site + ' tracker.')
            continue

        # Find the 'Date' row
        r = 1
        while ws.cell(row=r, column=date_col_num).value != 'Date':
            r += 1
            if r == 6:
                next_sheet = True
                break
        if next_sheet:
            continue

        # Extract date and uvs until the 'Total' row
        r += 1
        while ((ws.cell(row=r, column=date_col_num).value != 'Total') &
                   (ws.cell(row=r, column=date_col_num + 1).value != 'Total') &
                   (ws.cell(row=r, column=date_col_num).value != 'Total MTD')):
            uv = ws.cell(row=r, column=uv_col_num).value
            uv_list.append([site, sheet, ws.cell(row=r, column=date_col_num).value, uv])
            r += 1

    uv_df = pd.DataFrame(uv_list, columns=['Site', 'Original Report Tab Name', 'Date', 'UVs'])

    # Pick only this month's uvs
    mo = mo_year[0]
    year = mo_year[1]
    month_start_date = date(year, mo, 1)
    month_end_date = start_end_month(month_start_date)[1]

    uv_df['Date'] = [(d.date()) for d in uv_df['Date']]
    uv_df = uv_df[(uv_df['Date'] >= month_start_date) & (uv_df['Date'] <= month_end_date)]

    return uv_df

def get_cc_uvs(site, mo_year, cpuv_goals_sheet):

    ###########################################################
    # Make a list of (google sheet range, col name)
    ###########################################################

    cpuv_goals = get_cpuv_goals(mo_year[1], cpuv_goals_sheet)
    temp_site_gsrange_col_list = cpuv_goals['CC Partner Report'].tolist()

    gsrange_col_list = []
    for i in range(len(temp_site_gsrange_col_list)):
        temp_site_gsrange_col = temp_site_gsrange_col_list[i]
        if (temp_site_gsrange_col is not None) & (temp_site_gsrange_col != ''):
            for site_gsrange_col in [t.strip() for t in temp_site_gsrange_col.split(',')]:
                if site_gsrange_col.startswith(site[0]):
                    gsrange_col_list.append(site_gsrange_col[2:].split('_'))

    ###########################################################
    # Get uvs
    ###########################################################

    service = get_gsheet_service()
    spreadsheetId = CC_TRACKER_GSHEET[site]

    output = pd.DataFrame()
    for gsrange, col in gsrange_col_list:
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheetId, range=gsrange, valueRenderOption='UNFORMATTED_VALUE').execute()
        values = result.get('values', [])

        df = pd.DataFrame(values)
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)

        df = df[['Date', col]].rename(columns={col: 'UVs'})
        df['Original Report Tab Name'] = site[0] + '_' + gsrange + '_' + col
        output = pd.concat([output, df])

    ###########################################################
    # Pick only this month's uvs
    ###########################################################

    mo = mo_year[0]
    year = mo_year[1]
    month_start_date = date(year, mo, 1)
    month_end_date = start_end_month(month_start_date)[1]

    output = output[[isinstance(d, int) for d in output['Date']]]
    output['Date'] = [date(1900, 1, 1) + timedelta(days=(int(d) - 2)) for d in output['Date']]
    output = output[(output['Date'] >= month_start_date) & (output['Date'] <= month_end_date)]

    ###########################################################
    # Pick only number values for uvs
    ###########################################################

    output = output[pd.notnull(output['UVs']) & [not isinstance(uv, str) for uv in output['UVs']]]
    output['UVs'] = [int(uv) for uv in output['UVs']]

    ###########################################################
    # Add site and order columns
    ###########################################################

    output['Site'] = site
    output = output[['Site', 'Original Report Tab Name', 'Date', 'UVs']]
    output = output.reset_index(drop=True)

    return output

#####################################################
# DCM Reports
# start_date, end_date in the form of 'yyyy-mm-dd'
#####################################################

def dcm_reporting(start_date, end_date):

    DIR_PICKLES = DIR_PICKLES_ADP
    check_and_make_dir(DIR_PICKLES)

    # Check if a pickle file exists.
    pickle_file_name = 'dcm_reports_' + start_date + '_' + end_date + '.pickle'
    if os.path.exists(DIR_PICKLES + '/' + pickle_file_name):
        with open(DIR_PICKLES + '/' + pickle_file_name, 'rb') as f:
            return pickle.load(f)

    ######################################################

    print(datetime.now())

    output_folder_name = 'DCM Reports'
    if os.path.exists(output_folder_name):
        shutil.rmtree(output_folder_name)
    os.makedirs(output_folder_name)

    ##0. Prep
    service = dfareporting_utils.setup(None)

    report = {
        'name': 'Test HL AdOps',
        'type': 'STANDARD',
        # 'format': 'EXCEL',
        'criteria': {
            'dateRange': {'startDate': start_date, 'endDate': end_date},
            'dimensions': [{'name': 'dfa:campaign'}, {'name': 'dfa:placement'}, {'name': 'dfa:placementId'}],
            'metricNames': ['dfa:impressions', 'dfa:clicks', 'dfa:clickRate',
                            'dfa:activeViewViewableImpressions', 'dfa:activeViewMeasurableImpressions',
                            'dfa:activeViewEligibleImpressions', 'dfa:totalConversions']
        }
    }

    report_no_conv = {
        'name': 'Test HL AdOps',
        'type': 'STANDARD',
        # 'format': 'EXCEL',
        'criteria': {
            'dateRange': {'startDate': start_date, 'endDate': end_date},
            'dimensions': [{'name': 'dfa:campaign'}, {'name': 'dfa:placement'}, {'name': 'dfa:placementId'}],
            'metricNames': ['dfa:impressions', 'dfa:clicks', 'dfa:clickRate',
                            'dfa:activeViewViewableImpressions', 'dfa:activeViewMeasurableImpressions',
                            'dfa:activeViewEligibleImpressions']
        }
    }

    ##1. Get profile id's
    print('Getting profile id\'s...')
    try:
        request = service.userProfiles().list()
        response = request.execute()

        for_df = {}
        for profile in response['items']:
            for_df[profile['profileId']] = profile
        profiles_df = pd.DataFrame(for_df).transpose()

    except client.AccessTokenRefreshError:
        print('The credentials have been revoked or expired, please re-run the application to re-authorize')

    ##2. Create a report in DCM
    print('Creating DCM reports...')
    for i, row in profiles_df.iterrows():
        profile_id = row['profileId']
        account_name = row['accountName']
        output_file_name = account_name + '_' + str(profile_id) + '_' + start_date + '_' + end_date
        output_file_name = output_file_name.replace(':', ' ')
        output_file_name = output_file_name.replace('/', '')
        output_file_name = output_file_name + '.csv'
        profiles_df.loc[profiles_df['profileId'] == profile_id, 'Output File Name'] = output_file_name

        try:
            request = service.reports().insert(profileId=profile_id, body=report)
            response = request.execute()
            report_id = response['id']
            profiles_df.loc[profiles_df['profileId'] == profile_id, 'report_id'] = report_id
            profiles_df.loc[profiles_df['profileId'] == profile_id, 'has_conversions'] = True

        except googleapiclient.errors.HttpError:
            request = service.reports().insert(profileId=profile_id, body=report_no_conv)
            response = request.execute()
            report_id = response['id']
            profiles_df.loc[profiles_df['profileId'] == profile_id, 'report_id'] = report_id
            profiles_df.loc[profiles_df['profileId'] == profile_id, 'has_conversions'] = False

        except client.AccessTokenRefreshError:
            print('The credentials have been revoked or expired, please re-run the application to re-authorize')

    ##3. Run a report
    print('Running reports...')
    for i, row in profiles_df.iterrows():
        profile_id = row['profileId']
        report_id = row['report_id']

        try:
            request = service.reports().run(profileId=profile_id, reportId=report_id)
            result = request.execute()
            file_id = result['id']
            profiles_df.loc[profiles_df['profileId'] == profile_id, 'file_id'] = file_id

        except client.AccessTokenRefreshError:
            print('The credentials have been revoked or expired, please re-run the application to re-authorize')

    ##4. Keep looping till all files are downloaded
    profiles_df['report_downloaded'] = False

    print('Downloading files...')
    j = 1
    while len(profiles_df[profiles_df['report_downloaded'] == False]) > 0:
        if j >= 20:
            print('Not al files were downloaded.')
            break

        time.sleep(5)
        print('Round ' + str(j))
        ##4.1 Get a report status
        for i, row in profiles_df.iterrows():
            report_downloaded = row['report_downloaded']
            profile_id = row['profileId']
            report_id = row['report_id']
            file_id = row['file_id']

            if not report_downloaded:
                try:
                    request = service.reports().files().list(profileId=profile_id, reportId=report_id)

                    while True:
                        response = request.execute()
                        for report_file in response['items']:
                            if report_file['id'] == file_id:
                                profiles_df.loc[profiles_df['profileId'] == profile_id, 'report_status'] = report_file[
                                    'status']
                                break
                        break

                except client.AccessTokenRefreshError:
                    print('The credentials have been revoked or expired, please re-run the application to re-authorize')

        ##4.2 Download a file and Delete a report
        for i, row in profiles_df.iterrows():
            report_status = row['report_status']
            report_downloaded = row['report_downloaded']

            if (report_status == 'REPORT_AVAILABLE') & (not report_downloaded):
                profile_id = row['profileId']
                report_id = row['report_id']
                file_id = row['file_id']
                output_file_name = row['Output File Name']
                account_name = row['accountName']

                ##4.2.1 Download a file
                try:
                    request = service.files().get_media(reportId=report_id, fileId=file_id)

                    f = open(os.path.join(output_folder_name, output_file_name), 'wb')
                    f.write(request.execute())
                    f.close()

                    profiles_df.loc[profiles_df['profileId'] == profile_id, 'report_downloaded'] = True

                except client.AccessTokenRefreshError:
                    print (
                    'The credentials have been revoked or expired, please re-run the application to re-authorize')

                ##4.2.2 Delete a report
                try:
                    request = service.reports().delete(profileId=profile_id, reportId=report_id)
                    request.execute()
                    print(account_name.encode('utf-8'))

                except client.AccessTokenRefreshError:
                    print('The credentials have been revoked or expired, please re-run the application to re-authorize')
        j += 1

    ##5. Delete files with no data, and create a file with all data
    print('Finishing up...')
    all_data = pd.DataFrame()

    for i, row in profiles_df.iterrows():
        output_file_name = row['Output File Name']
        if row['report_downloaded']:
            account_id = row['accountId']

            f = open(output_folder_name + '/' + output_file_name, encoding='utf8')
            r = 0
            for line in csv.reader(f):
                if len(line) > 0:
                    if line[0] == 'Campaign':
                        f.close()
                        break
                    else:
                        r += 1
                else:
                    r += 1
            data = pd.read_csv(output_folder_name + '/' + output_file_name, skiprows=r, skipfooter=1, engine='python',
                               encoding='utf-8')
            profiles_df.loc[profiles_df['Output File Name'] == output_file_name, '# of Data Rows'] = len(data)

            if len(data) == 0:
                os.remove(output_folder_name + '/' + output_file_name)
            else:
                data['Account ID'] = int(account_id)
                all_data = all_data.append(data)
        else:
            profiles_df.loc[profiles_df['Output File Name'] == output_file_name, 'Output File Name'] = ''

    all_data = all_data.sort_values(['Campaign', 'Placement'])[
        ['Account ID', 'Campaign', 'Placement', 'Placement ID', 'Impressions', 'Clicks',
         'Click Rate', 'Active View: Viewable Impressions',
         'Active View: Measurable Impressions', 'Active View: Eligible Impressions',
         'Total Conversions']]

    # Dedupe all_data
    all_data = all_data.drop_duplicates()

    profiles_df = profiles_df[['accountId', 'accountName', 'profileId', 'subAccountId', 'subAccountName',
                               'userName', 'has_conversions', 'report_downloaded', '# of Data Rows',
                               'Output File Name']].sort_values('Output File Name')
    for col in ['accountId', 'profileId', 'subAccountId']:
        profiles_df[col] = [int(entry) if isinstance(entry, str) else entry for entry in profiles_df[col]]

    ##6. Group data by placement name without size
    grouped_all_data = all_data.copy()
    grouped_all_data['Placement w/o Space'] = [p.replace(' ', '') for p in grouped_all_data['Placement']]

    size_pattern = re.compile('[0-9]+x[0-9]+')

    def remove_size(placement):
        size_list = size_pattern.findall(placement)
        if len(size_list) > 0:
            output = placement
            for size in size_list:
                output = output.replace(size, '')
            return output
        else:
            return placement

    grouped_all_data['Placement w/o Size'] = [remove_size(p) for p in grouped_all_data['Placement w/o Space']]

    grouped_all_data = grouped_all_data[['Account ID', 'Campaign', 'Placement w/o Size',
                                         'Impressions', 'Clicks', 'Click Rate', 'Active View: Viewable Impressions',
                                         'Active View: Measurable Impressions', 'Active View: Eligible Impressions',
                                         'Total Conversions']]

    grouped_all_data = grouped_all_data.groupby(['Account ID', 'Campaign', 'Placement w/o Size']).sum().reset_index()
    grouped_all_data['Click Rate'] = grouped_all_data['Clicks'] / grouped_all_data['Impressions'] * 100

    print(datetime.now())

    # Save to pickle
    with open(DIR_PICKLES + '/' + pickle_file_name, 'wb') as f:
        pickle.dump((all_data, profiles_df, grouped_all_data), f)

    return (all_data, profiles_df, grouped_all_data)


"""
##Delete Test Reports
for i, row in profiles_df.iterrows():
    profile_id = row['profileId']

    request = service.reports().list(profileId=profile_id)

    while True:
        response = request.execute()
        for report in response['items']:
            print(report['name'])
            if report['name'] == 'Test HL AdOps':
                report_id = report['id']
                request = service.reports().delete(profileId=profile_id, reportId=report_id)
                request.execute()
                print('successfully deleted')
        break
"""


#####################################################
# Allergan Reports
# start_date, end_date in the form of 'yyyymmdd'
#####################################################

def allergan_report(start_date, end_date):
    DIR_PICKLES = DIR_PICKLES_ADP
    check_and_make_dir(DIR_PICKLES)

    # Check if a pickle file exists.
    pickle_file_name = 'allergan_report_' + start_date + '_' + end_date + '.pickle'
    if os.path.exists(DIR_PICKLES + '/' + pickle_file_name):
        with open(DIR_PICKLES + '/' + pickle_file_name, 'rb') as f:
            return pickle.load(f)

    ############################################
    # Get reports from the web
    ############################################

    s = requests.Session()
    r = s.get('https://mediacomallergan.moat.com/users/login')
    login_data = {'data[User][email]': 'healthline_allergan@moat.com', 'data[User][password]': 'f4Wy8^T3'}
    r2 = s.post('https://mediacomallergan.moat.com/users/login', data=login_data)
    query_url_1 = 'https://mediacomallergan.moat.com/exports/export_data?brandId=27uz&startDate=' + start_date + '&endDate=' + end_date + '&columns=level1%2Clevel3%2Cloads%2Cin_vp_meas%2Chuman_and_viewable%2Chuman_and_groupm_payable_sum&destination=%5B%5D'
    query_url_2 = 'https://mediacomallergan.moat.com/exports/export_data?brandId=Buq&startDate=' + start_date + '&endDate=' + end_date + '&columns=level1%2Clevel3%2Cloads%2Cin_vp_meas%2Chuman_and_viewable%2Chuman_and_groupm_payable_sum&destination=%5B%5D'

    r3 = s.get(query_url_1)
    f = open('moat_allergan_report_mediacom.csv', 'w')
    f.write(r3.text)
    f.close()

    #r4 = s.get(query_url_2)
    #f = open('moat_allergan_report_medialets.csv', 'w')
    #f.write(r4.text)
    #f.close()

    moat_report_1 = pd.read_csv('moat_allergan_report_mediacom.csv')
    moat_report_1['Server'] = 'Mediacom'

    #moat_report_2 = pd.read_csv('moat_allergan_report_medialets.csv')
    #moat_report_2['Server'] = 'Medialets'

    #moat_report = pd.concat([moat_report_1, moat_report_2])
    moat_report = moat_report_1.copy()
    o1_original = moat_report.copy()

    os.remove('moat_allergan_report_mediacom.csv')
    #os.remove('moat_allergan_report_medialets.csv')

    ############################################
    # Extract some info
    ############################################

    # Add Site
    moat_report['Site'] = ''

    moat_report.loc[pd.isnull(moat_report['Placement Label']), 'Placement Label'] = 'N/A'
    moat_report.loc[moat_report['Placement Label'] == 'N/A', 'Site'] = 'N/A'
    moat_report.loc[moat_report['Placement Label'].str.contains('drugs.com', case=False), 'Site'] = 'Drugs'
    moat_report.loc[moat_report['Placement Label'].str.contains('av', case=False), 'Site'] = 'HL - AV'
    moat_report.loc[moat_report['Site'] == '', 'Site'] = 'HL'

    # Add Device
    moat_report['Device'] = ''
    moat_report.loc[moat_report['Placement Label'].str.contains('_De_', case=False), 'Device'] = 'Desktop'
    moat_report.loc[moat_report['Placement Label'].str.contains('_Ta_', case=False), 'Device'] = 'Tablet'
    moat_report.loc[moat_report['Placement Label'].str.contains('_Mo_', case=False), 'Device'] = 'Mobile'

    # Add Advertiser
    dict_advertiser = {'AG_2016_FY_Botox CM Branded Display': 'Others',
                       'AG_2016_FY_Botox CM Branded Display ': 'Others',
                       'AG_2016_FY_Linzess': 'Others',
                       'AG_2016_FY_RESBranded': 'Others',
                       'AG_2016_FY_Viberzi': 'Others',
                       'AG_2017_FY_Botox CM Branded Display': 'Botox',
                       'AG_2017_FY_Botox_CM_Branded_Display': 'Botox',
                       'AG_2017_FY_Linzess Display': 'Linzess',
                       'AG_2017_FY_Namzaric Branded Display': 'Namzaric',
                       'AG_2017_FY_Restasis Branded Display': 'Restasis',
                       'AG_2017_FY_Viberzi Display': 'Viberzi',
                       'none': 'Others'}

    moat_report['Campaign Label'] = [(campaign_label if isinstance(campaign_label, str) else 'none') for campaign_label
                                     in moat_report['Campaign Label']]
    moat_report['Advertiser'] = [dict_advertiser[campaign_label] for campaign_label in moat_report['Campaign Label']]

    # Exclude non-CPM placements!!!
    moat_report.loc[
        (moat_report['Advertiser'] == 'Restasis') & moat_report['Placement Label'].str.contains('Custom Sponsorship',
                                                                                                False), 'Advertiser'] = 'CPUV Restasis'
    moat_report.loc[
        (moat_report['Advertiser'] == 'Restasis') & moat_report['Placement Label'].str.contains('Topic Center',
                                                                                                False), 'Advertiser'] = 'CPUV Restasis'
    moat_report.loc[(moat_report['Advertiser'] == 'Linzess') & moat_report['Placement Label'].str.contains(
        'Competitive Conquesting', False), 'Advertiser'] = 'CPUV Linzess'
    moat_report.loc[(moat_report['Advertiser'] == 'Linzess') & moat_report['Placement Label'].str.contains(
        'Chronic Constipation Center', False), 'Advertiser'] = 'CPUV Linzess'

    # Add Size
    size_pattern = re.compile('_([0-9]+\sx\s[0-9]+)_')
    moat_report['Size'] = [
        'N/A' if not size_pattern.search(placement_label) else size_pattern.search(placement_label).group(1) for
        placement_label in moat_report['Placement Label']]

    # Save the labeled Moat report
    o2_labeled = moat_report.copy()

    ############################################
    # Caculate billable imps
    ############################################

    # Remove Non-live campaigns
    moat_report = moat_report.loc[(moat_report['Advertiser'] == 'Botox') |
                                  (moat_report['Advertiser'] == 'Linzess') |
                                  (moat_report['Advertiser'] == 'Namzaric') |
                                  (moat_report['Advertiser'] == 'Restasis') |
                                  (moat_report['Advertiser'] == 'Viberzi')]

    # Remove Botox CM, Tablet and Mobile, Mediacom
    #moat_report = moat_report[-((moat_report['Advertiser'] == 'Botox') &
    #                            ((moat_report['Device'] == 'Tablet') | (moat_report['Device'] == 'Mobile')) &
    #                            (moat_report['Server'] == 'Mediacom'))]

    # Create a pivot table
    moat_report = moat_report[(moat_report['Site'] == 'Drugs') | (moat_report['Site'] == 'HL')]
    moat_report = moat_report[['Site', 'Advertiser', 'Device', 'Size', 'Impressions Analyzed',
                               'On-Screen Measurable Impressions', 'Human and Viewable Impressions',
                               'GroupM Payable Impressions']]
    pivot = moat_report.groupby(['Site', 'Advertiser', 'Device', 'Size']).sum().reset_index()

    # Add Total for each (Site, Advertiser, Device)
    pivot_by_site_adv_device = moat_report[['Site', 'Advertiser', 'Device', 'Impressions Analyzed',
                                            'On-Screen Measurable Impressions', 'Human and Viewable Impressions',
                                            'GroupM Payable Impressions']].groupby(
        ['Site', 'Advertiser', 'Device']).sum().reset_index()
    pivot_by_site_adv_device['Size'] = 'Total'
    pivot = pd.concat([pivot, pivot_by_site_adv_device])

    # Add Total for each (Site, Advertiser)
    pivot_by_site_adv = moat_report[['Site', 'Advertiser', 'Impressions Analyzed',
                                     'On-Screen Measurable Impressions', 'Human and Viewable Impressions',
                                     'GroupM Payable Impressions']].groupby(['Site', 'Advertiser']).sum().reset_index()
    pivot_by_site_adv['Device'] = 'Total'
    pivot = pd.concat([pivot, pivot_by_site_adv])

    # Add Viewability
    pivot['Viewability'] = pivot['Human and Viewable Impressions'] / pivot['On-Screen Measurable Impressions']

    # Add Viewability To Apply for Mobile
    ## Desktop 300x250 Viewability for Mobile 300x250
    ## Desktop Overall Viewability for Mobile 320x50 and 300x50
    pivot['Viewability To Apply'] = ''
    list_of_list_site_adv_dev_size = pivot[['Site', 'Advertiser', 'Device', 'Size']].drop_duplicates().values.tolist()
    for site, adv, dev, size in list_of_list_site_adv_dev_size:
        if (dev == 'Mobile') & (size != 'Total'):
            if size == '300 x 250':
                use_size = '300 x 250'
            else:
                use_size = 'Total'

            use_df = pivot[(pivot['Site'] == site) &
                           (pivot['Advertiser'] == adv) &
                           (pivot['Device'] == 'Desktop') &
                           (pivot['Size'] == use_size)]

            if len(use_df) > 0:
                use_viewability = use_df['Viewability'].tolist()[0]
            else:
                use_viewability = 0.0

            pivot.loc[(pivot['Site'] == site) &
                      (pivot['Advertiser'] == adv) &
                      (pivot['Device'] == dev) &
                      (pivot['Size'] == size), 'Viewability To Apply'] = use_viewability

    # Add Billable Imps
    pivot['Billable Imps'] = pivot['GroupM Payable Impressions']

    ## Mobile Billable Imps
    pivot.loc[pivot['Device'] == 'Mobile', 'Billable Imps'] = pivot.loc[pivot['Device'] == 'Mobile', 'Viewability To Apply'] * pivot.loc[pivot['Device'] == 'Mobile', 'Impressions Analyzed']
    for site, adv, dev, size in list_of_list_site_adv_dev_size:
        if (dev == 'Mobile') & (size == 'Total'):
            total_billable = pivot[(pivot['Site'] == site) &
                                   (pivot['Advertiser'] == adv) &
                                   (pivot['Device'] == 'Mobile') &
                                   (pivot['Size'] != 'Total')]['Billable Imps'].sum()
            pivot.loc[(pivot['Site'] == site) &
                      (pivot['Advertiser'] == adv) &
                      (pivot['Device'] == 'Mobile') &
                      (pivot['Size'] == 'Total'), 'Billable Imps'] = total_billable

    ## Overall Billable Imps
    for site, adv, dev, size in list_of_list_site_adv_dev_size:
        if dev == 'Total':
            total_billable = pivot[(pivot['Site'] == site) &
                                   (pivot['Advertiser'] == adv) &
                                   (pivot['Size'] == 'Total')]['Billable Imps'].sum()
            pivot.loc[(pivot['Site'] == site) &
                      (pivot['Advertiser'] == adv) &
                      (pivot['Device'] == 'Total'), 'Billable Imps'] = total_billable

    # Sort
    pivot['DFP Imps'] = ''
    pivot['Billability'] = ''
    pivot['Discrepancy'] = ''

    pivot = pivot.sort_values(['Site', 'Advertiser', 'Device', 'Size'])
    pivot = pivot[['Site', 'Advertiser', 'Device', 'Size', 'Human and Viewable Impressions',
                   'On-Screen Measurable Impressions', 'Viewability', 'Viewability To Apply',
                   'Impressions Analyzed', 'Billable Imps', 'DFP Imps', 'Billability', 'Discrepancy',
                   'GroupM Payable Impressions']]

    o3_summary = pivot.copy()
    o4_only_billable = pivot[pivot['Device'] == 'Total'][['Advertiser', 'Site', 'Billable Imps']]

    ############################################
    # Save to pickle
    ############################################

    with open(DIR_PICKLES + '/' + pickle_file_name, 'wb') as f:
        pickle.dump((o1_original, o2_labeled, o3_summary, o4_only_billable), f)

    return (o1_original, o2_labeled, o3_summary, o4_only_billable)

#####################################################
# Get DFP 'DCM Placement IDs' report
# Used for linking DFP and DCM
# Given 'end_date', it outputs MTD report
#####################################################

def get_dcm_placement_ids(end_date):

    DIR_PICKLES = DIR_PICKLES_ADP
    check_and_make_dir(DIR_PICKLES)

    # Prep
    start_date = end_date.replace(day=1)
    str_start_date = str(start_date.year) + str(start_date.month).zfill(2) + str(start_date.day).zfill(2)
    str_end_date = str(end_date.year) + str(end_date.month).zfill(2) + str(end_date.day).zfill(2)

    # Check if a pickle file exists.
    pickle_file_name = 'dcm_placement_ids_' + str_start_date + '_' + str_end_date + '.pickle'
    if os.path.exists(DIR_PICKLES + '/' + pickle_file_name):
        with open(DIR_PICKLES + '/' + pickle_file_name, 'rb') as f:
            return pickle.load(f)

    # CVS file name
    output_file_name = 'DCM_Placement_Ids_' + str_start_date + '_' + str_end_date + '.csv'

    # Report query id
    saved_query_id = '10002557419'

    # Initialize a client
    client = dfp.DfpClient.LoadFromStorage()

    # Initialize appropriate service.
    report_service = client.GetService('ReportService', version='v201705')

    # Initialize a DataDownloader.
    report_downloader = client.GetDataDownloader(version='v201705')

    # Create statement object to filter for an order.
    values = [{'key': 'id',
               'value': {'xsi_type': 'NumberValue',
                         'value': saved_query_id}}]
    query = 'WHERE id = :id'
    statement = dfp.FilterStatement(query, values, 1)

    response = report_service.getSavedQueriesByStatement(statement.ToStatement())

    if 'results' in response:
        saved_query = response['results'][0]

        if saved_query['isCompatibleWithApiVersion']:
            report_job = {}

            # Set report query and optionally modify it.
            report_job['reportQuery'] = saved_query['reportQuery']

            report_job['reportQuery']['startDate']['year'] = start_date.year
            report_job['reportQuery']['startDate']['month'] = start_date.month
            report_job['reportQuery']['startDate']['day'] = start_date.day

            report_job['reportQuery']['endDate']['year'] = end_date.year
            report_job['reportQuery']['endDate']['month'] = end_date.month
            report_job['reportQuery']['endDate']['day'] = end_date.day

            try:
                # Run the report and wait for it to finish.
                report_job_id = report_downloader.WaitForReport(report_job)
            except errors.DfpReportError as e:
                print('Failed to generate report. Error was: %s' % e)

            # Download report data.
            with open(output_file_name, 'wb') as report_file:
                report_downloader.DownloadReportToFile(report_job_id, 'CSV_DUMP', report_file, use_gzip_compression=False)

        else:
            print('The query specified is not compatible with the API version.')

    # Clean up
    cols = ['Dimension.ORDER_NAME',
            'Dimension.LINE_ITEM_NAME',
            'Dimension.CREATIVE_NAME',
            'CF[6995]_Value',
            'CF[7115]_Value',
            'Column.TOTAL_LINE_ITEM_LEVEL_IMPRESSIONS']

    col_rename_dict = {'Dimension.ORDER_NAME': 'Order',
                       'Dimension.LINE_ITEM_NAME': 'Line item',
                       'Dimension.CREATIVE_NAME': 'Creative',
                       'CF[6995]_Value': 'DAS Line Item Name',
                       'CF[7115]_Value': '3rd Party Creative ID',
                       'Column.TOTAL_LINE_ITEM_LEVEL_IMPRESSIONS': 'Total impressions'}

    df = pd.read_csv(output_file_name, encoding='utf-8')
    df = df[cols].rename(columns=col_rename_dict)

    # Write to a pickle file
    with open(DIR_PICKLES + '/' + pickle_file_name, 'wb') as f:
        pickle.dump(df, f)

    # Delete the csv file
    os.remove(output_file_name)

    return df

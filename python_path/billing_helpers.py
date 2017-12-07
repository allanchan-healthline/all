from path2pickles import *
from NEW_helpers import *
from delivery_helpers import *
from always_up2date_helpers import *
from gsheet_gdrive_api import *

import pickle
import pandas as pd
import openpyxl as opx

from datetime import date

pd.options.mode.chained_assignment = None


##############################################################
# Billing Grid
##############################################################

def make_bg_as_csv(year, mo):

    ##############################################################
    # Prep
    ##############################################################

    output_file_name = 'for_billing_grid_' + str(year) + '_' + str(mo).zfill(2) + '.csv'

    start_date = date(year, mo, 1)
    end_date = start_end_month(start_date)[1]

    def convert2dcm_date(date_obj):
        return str(date_obj.year) + '-' + str(date_obj.month).zfill(2) + '-' + str(date_obj.day).zfill(2)

    start_date = convert2dcm_date(start_date)
    end_date = convert2dcm_date(end_date)

    ##############################################################
    # CPM DFP imps
    ##############################################################
        
    DIR_PICKLES = PATH2PICKLES + '/' + PREFIX4ALWAYS_UP2DATE + str(year) + '_' + str(mo).zfill(2)
    with open(DIR_PICKLES + '/' + 'all1.pickle', 'rb') as f:
        all1 = pickle.load(f)
    
    all1_cpm = all1[all1['Price Calculation Type'] == 'CPM'].rename(columns={'Impressions/UVs': 'First Party Units'})
    
    bbr_brand_li = ['(DAS)BBR #', 'Brand', 'DAS Line Item Name']
    dfp_imps = all1_cpm[bbr_brand_li + ['First Party Units']].groupby(bbr_brand_li).sum().reset_index()

    ##############################################################
    # DCM imps
    ##############################################################

    all_data, profiles_df, grouped_all_data = dcm_reporting(start_date, end_date)

    groupby_col = ['Placement ID']
    value_col = ['Impressions']
    rename_dict = {'Placement ID': '3rd Party Creative ID', 'Impressions': 'DFA(by ID)'}
    dcm_imps = all_data[groupby_col + value_col].groupby(groupby_col).sum().reset_index().rename(columns=rename_dict)
    dcm_imps['3rd Party Creative ID'] = [str(id) for id in dcm_imps['3rd Party Creative ID']]
    
    dfp_dcm_map = all1_cpm[bbr_brand_li + ['3rd Party Creative ID']].drop_duplicates()
    dcm_imps = pd.merge(dfp_dcm_map, dcm_imps, how='left', on='3rd Party Creative ID')
    dcm_imps = dcm_imps[bbr_brand_li + ['DFA(by ID)']].groupby(bbr_brand_li).sum().reset_index()

    ##############################################################
    # Select Billing Grid data from DAS
    ##############################################################

    das = make_das(use_scheduled_units=False, export=False)
    das_month = str(mo) + '/' + str(year)
    
    bg = das[das[das_month] > 0]
    bg = bg[bg['Price Calculation Type'] != 'CPA']

    ##############################################################
    # Add Expedited Invoice
    ##############################################################

    exp_inv = get_expedited_invoice_opportunities()
    exp_inv = exp_inv[['BBR', 'Expedited Invoice']].drop_duplicates()
    bg = pd.merge(bg, exp_inv, how='left', on=['BBR'])

    ##############################################################
    # Multi-Month Bill Up To
    ##############################################################
    
    das_header = das.columns.tolist()
    months_list = []
    for col in das_header:
        if re.search('[0-9]+/[0-9]{4}', col):
            months_list.append(col)

    (das_month_mo, das_month_year) = das_month.split('/')
    das_month_mo = int(das_month_mo)
    das_month_year = int(das_month_year)

    months = []
    for month in months_list:
        (mo, year) = month.split('/')
        mo = int(mo)
        year = int(year)
        if year < das_month_year:
            continue
        elif year > das_month_year:
            months.append(month)
        elif mo >= das_month_mo:
                months.append(month)

    for month in months:
        bg[month].fillna(0, inplace=True)

    bg['Multi-Month Bill Up To'] = bg.apply(lambda row: sum(round(row[month], 0) for month in months), axis=1)
    bg.loc[bg['Flight Type'] == 'Monthly', 'Multi-Month Bill Up To'] = ''

    ##############################################################
    # Rename columns
    ##############################################################

    rename_dict = {'Brand': 'Advertiser', 
                   'IO Number': 'Purchase Order / Insertion Order',
                   'Line Description': 'Placement', 
                   'Price Calculation Type': 'Rate Type', 
                   'Sales Price': 'Rate', 
                   das_month: 'Booked Impressions', 
                   'Flight Type': 'Goal Breakdown', 
                   'Account Manager': 'AM', 
                   'Campaign Manager': 'CM',
                   'Billable Reporting Source': 'Third Party System'}
    
    bg = bg.rename(columns=rename_dict)

    ##############################################################
    # Add CPM DFP and DCM imps
    ##############################################################

    imps_rename_dict = {'(DAS)BBR #': 'BBR', 'Brand': 'Advertiser', 'DAS Line Item Name': 'Placement'}
    imps_join_on = ['BBR', 'Advertiser', 'Placement']

    bg = pd.merge(bg, dfp_imps.rename(columns=imps_rename_dict), how='left', on=imps_join_on)
    bg = pd.merge(bg, dcm_imps.rename(columns=imps_rename_dict), how='left', on=imps_join_on)

    ##############################################################
    # Add columns to be manually entered or calculated with formulas
    ##############################################################

    bg['Billed Units'] = ''
    bg['Total Cost'] = ''
    bg['110%'] = ''
    bg['Discrepancy'] = ''
    bg['Third Party Impressions'] = ''
    bg['Check DFA'] = ''
    bg['Overbilling Check'] = ''
    bg['Delivery to Booked Check'] = ''
    bg['DAS Cost'] = ''
    bg['Actual Cost'] = ''
    bg['DAS v Actual Cost'] = ''
    bg['Hit the goal?'] = ''
    bg['UD'] = ''
    bg['UD $'] = ''
    bg['Confirmed?'] = 0

    ##############################################################
    # Sort
    ##############################################################

    header = ['OLI', 'Advertiser', 'Agency', 'Campaign Name', 'Purchase Order / Insertion Order', 'Line Item Number',
              'Placement', 'Rate Type', 'Rate', 'Billed Units', 'Total Cost', 'Booked Impressions', '110%', 'Multi-Month Bill Up To',
              'Discrepancy', 'First Party Units', 'Third Party Impressions', 'DFA(by ID)', 'Check DFA', 'Third Party System',
              'Goal Breakdown', 'AM', 'CM', 'BBR', 'Expedited Invoice', 'DAS Cost', 'Actual Cost',
              'DAS v Actual Cost', 'Hit the goal?', 'UD', 'UD $', 'Confirmed?', 'Stage']

    sortby = ['Advertiser', 'Campaign Name', 'BBR', 'Line Item Number', 'Placement']

    bg = bg[header].sort_values(sortby).reset_index(drop=True)

    ##############################################################
    # Formulas
    ##############################################################

    def col_name2col_letter(col_name):
        col_num = header.index(col_name) + 1
        return opx.utils.get_column_letter(col_num)

    col_third_party_system = col_name2col_letter('Third Party System')
    col_goal_breakdown = col_name2col_letter('Goal Breakdown')
    col_booked_imps = col_name2col_letter('Booked Impressions')
    col_first_party_units = col_name2col_letter('First Party Units')
    col_110 = col_name2col_letter('110%')
    col_multimonth_billupto = col_name2col_letter('Multi-Month Bill Up To')
    col_third_party_imps = col_name2col_letter('Third Party Impressions')
    col_rate_type = col_name2col_letter('Rate Type')
    col_billed_units = col_name2col_letter('Billed Units')
    col_rate = col_name2col_letter('Rate')
    col_total_cost = col_name2col_letter('Total Cost')
    col_actual_cost = col_name2col_letter('Actual Cost')
    col_das_cost = col_name2col_letter('DAS Cost')
    col_ud = col_name2col_letter('UD')
    col_dfa_byid = col_name2col_letter('DFA(by ID)')
    col_oli = col_name2col_letter('OLI')

    def get_billed_units(i):
        row = str(i + 2)
        output = '=ROUND(IF(OR(' + col_third_party_system + row + '="DFP", ' + col_third_party_system + row + '="The Trade Desk"), '
        output += 'IF(' + col_goal_breakdown + row + '="Monthly", MIN(' + col_booked_imps + row + ', ' + col_first_party_units + row + '), '
        output += 'MIN(' + col_110 + row + ', ' + col_multimonth_billupto + row + ', ' + col_first_party_units + row + ')), '
        output += 'IF(' + col_goal_breakdown + row + '="Monthly", MIN(' + col_booked_imps + row + ', ' + col_third_party_imps + row + '), '
        output += 'MIN(' + col_110 + row + ', ' + col_multimonth_billupto + row + ', ' + col_third_party_imps + row + '))), 0)'
        return output

    def get_total_cost(i):
        row = str(i + 2)
        output = '=IF(OR(' + col_rate_type + row + '="CPM", ' + col_rate_type + row + '="AV-CPM"), '
        output += col_billed_units + row + '/1000*' + col_rate + row + ', '
        output += col_billed_units + row + '*' + col_rate + row + ')'
        return output

    def get_110(i):
        row = str(i + 2)
        return '=IF(' + col_goal_breakdown + row + '<>"Monthly", ' + col_booked_imps + row + '*1.1, "")'

    def get_das_cost(i):
        row = str(i + 2)
        output = '=IF(OR(' + col_rate_type + row + '="CPM", ' + col_rate_type + row + '="AV-CPM"), '
        output += col_booked_imps + row + '/1000*' + col_rate + row + ', '
        output += col_booked_imps + row + '*' + col_rate + row + ')'
        return output

    def get_actual_cost(i):
        row = str(i + 2)
        return '=' + col_total_cost + row

    def get_dasvactual_cost(i):
        row = str(i + 2)
        return '=' + col_actual_cost + row + '-' + col_das_cost + row

    def get_hit_goal(i):
        row = str(i + 2)
        return '=IF(' + col_billed_units + row + '>=' + col_booked_imps + row + ', 1, 0)'

    def get_ud(i):
        row = str(i + 2)
        output = '=IF(' + col_booked_imps + row + '-' + col_billed_units + row + '>0, '
        output += col_booked_imps + row + '-' + col_billed_units + row + ', 0)'
        return output

    def get_ud_amount(i):
        row = str(i + 2)
        output = '=IF(OR(' + col_rate_type + row + '="CPM", ' + col_rate_type + row + '="AV-CPM"), '
        output += col_ud + row + '/1000*' + col_rate + row + ', '
        output += col_ud + row + '*' + col_rate + row + ')'
        return output

    def get_discrepancy(i):
        row = str(i + 2)
        output = '=IF(OR(' + col_third_party_system + row + '="DFP", ' + col_third_party_system + row + '="The Trade Desk"), 0, '
        output += '(' + col_first_party_units + row + '-' + col_third_party_imps + row + ')/' + col_first_party_units + row + ')'
        return output

    def get_check_dfa(i):
        row = str(i + 2)
        output = '=IF(ISNUMBER(FIND("DCM", ' + col_third_party_system + row + ')), '
        output += col_third_party_imps + row + '=' + col_dfa_byid + row + ', "")'
        return output

    def update_billed_units(row):
        if row['Rate Type'] == 'Flat-fee':
            if row['Stage'] == 'Booked Not Live':
                return 0
            return row['Booked Impressions']
        return row['Billed Units']

    bg['Billed Units'] = [get_billed_units(i) for i in range(len(bg))]
    bg['Total Cost'] = [get_total_cost(i) for i in range(len(bg))]
    bg['110%'] = [get_110(i) for i in range(len(bg))]
    bg['DAS Cost'] = [get_das_cost(i) for i in range(len(bg))]
    bg['Actual Cost'] = [get_actual_cost(i) for i in range(len(bg))]
    bg['DAS v Actual Cost'] = [get_dasvactual_cost(i) for i in range(len(bg))]
    bg['Hit the goal?'] = [get_hit_goal(i) for i in range(len(bg))]
    bg['UD'] = [get_ud(i) for i in range(len(bg))]
    bg['UD $'] = [get_ud_amount(i) for i in range(len(bg))]
    bg['Discrepancy'] = [get_discrepancy(i) for i in range(len(bg))]
    bg['Check DFA'] = [get_check_dfa(i) for i in range(len(bg))]

    # Not same for all rows
    bg.loc[bg['Rate Type'] != 'CPM', ('Discrepancy', 'Check DFA')] = ('', '')
    bg['Billed Units'] = bg.apply(lambda row: update_billed_units(row), axis=1)

    ##############################################################
    # Grand Total row
    ##############################################################

    last_row = str(len(bg) + 1)
    total_row_df = pd.DataFrame(index=[0], columns=bg.columns.tolist())

    # Subtotal
    subtotal_col_name_list = ['Billed Units',
                              'Total Cost',
                              'Booked Impressions',
                              '110%',
                              'Multi-Month Bill Up To',
                              'First Party Units',
                              'Third Party Impressions',
                              'DFA(by ID)',
                              'DAS Cost',
                              'Actual Cost',
                              'DAS v Actual Cost',
                              'Hit the goal?',
                              'UD',
                              'UD $',
                              'Confirmed?']

    for col_name in subtotal_col_name_list:
        col_letter = col_name2col_letter(col_name)
        total_row_df[col_name] = '=SUBTOTAL(9, ' + col_letter + '2:' + col_letter + last_row + ')'

    # Other
    total_row_df['OLI'] = 'Grand Total'
    total_row_df['Discrepancy'] = '=(' + col_first_party_units + last_row + '-' + col_third_party_imps + last_row + ')/' + col_first_party_units + last_row
    total_row_df['Check DFA'] = '=' + col_third_party_imps + last_row + '=' + col_dfa_byid + last_row

    ##############################################################
    # Output
    ##############################################################

    bg = pd.concat([bg, total_row_df])
    bg.to_csv(output_file_name, index=False)

    return output_file_name

def upload_bg2gdrive(year, mo, csv_file_name):

    #################################################
    folder_id = '0B71ox_2Qc7gmWkNWM21YUU9MSEU'
    #################################################

    service = get_gsheet_service()    
    
    # Upload csv to Google Drive as Google Sheet
    month_dict = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}
    gsheet_file_name = month_dict[mo] + ' ' + str(year) + ' Billing Grid'

    file_id = save_csv_as_gsheet_in_gdrive(gsheet_file_name, folder_id, csv_file_name)

    # Change sheet name to BG
    ss_metadata = service.spreadsheets().get(spreadsheetId=file_id).execute()
    s_metadata = ss_metadata['sheets'][0]['properties']
    sheet_id = s_metadata['sheetId']

    change_sheet_name = [{'updateSheetProperties': {'properties': {'sheetId': sheet_id, 'title': 'BG'}, 'fields': 'title'}}]
    result = service.spreadsheets().batchUpdate(spreadsheetId=file_id, body={'requests': change_sheet_name}).execute()

    # Add template BG formatting sheet
    #################################################
    template_ss_id = '1EM3EDQayLgn9MBsmdRnHhFxyq6553Q5waWTzpdzUogE'
    template_s_id = '1207419624'
    #################################################

    request = service.spreadsheets().sheets().copyTo(spreadsheetId=template_ss_id, sheetId=template_s_id, body={'destination_spreadsheet_id': file_id})
    response = request.execute()
    
    return file_id

def format_bg(file_id):

    service = get_gsheet_service()
    
    # Get sheet ids
    s_id_bg = None
    s_id_template = None
    
    ss_metadata = service.spreadsheets().get(spreadsheetId=file_id).execute()
    for s_metadata in ss_metadata['sheets']:
        if s_metadata['properties']['title'] == 'BG':
            s_id_bg = s_metadata['properties']['sheetId']
        if s_metadata['properties']['title'] == 'Copy of template':
            s_id_template = s_metadata['properties']['sheetId']

    # Get values > header, # of rows, # of columns
    result = service.spreadsheets().values().get(
        spreadsheetId=file_id, range='BG', valueRenderOption='UNFORMATTED_VALUE').execute()
    values = result.get('values', [])
    
    header = values[0]
    n_col = len(header)
    n_row = len(values)

    # Header formatting from template
    header_format = [{'copyPaste': {'source': {'sheetId': s_id_template,
                                               'startRowIndex': 0,
                                               'endRowIndex': 1,
                                               'startColumnIndex': 0,
                                               'endColumnIndex': n_col},
                                    'destination': {'sheetId': s_id_bg,
                                                    'startRowIndex': 0,
                                                    'endRowIndex': 1,
                                                    'startColumnIndex': 0,
                                                    'endColumnIndex': n_col},
                                    'pasteType': 'PASTE_FORMAT'}}]


    # Content formatting from template
    content_format = [{'copyPaste': {'source': {'sheetId': s_id_template, 
                                                'startRowIndex': 1,
                                                'endRowIndex': 2,
                                                'startColumnIndex': 0,
                                                'endColumnIndex': n_col},
                                    'destination': {'sheetId': s_id_bg,
                                                    'startRowIndex': 1,
                                                    'endRowIndex': n_row-1,
                                                    'startColumnIndex': 0,
                                                    'endColumnIndex': n_col},
                                    'pasteType': 'PASTE_FORMAT'}}]

    # Footer formatting from template
    footer_format = [{'copyPaste': {'source': {'sheetId': s_id_template,
                                               'startRowIndex': 2,
                                               'endRowIndex': 3,
                                               'startColumnIndex': 0,
                                               'endColumnIndex': n_col},
                                    'destination': {'sheetId': s_id_bg,
                                                    'startRowIndex': n_row-1,
                                                    'endRowIndex': n_row,
                                                    'startColumnIndex': 0,
                                                    'endColumnIndex': n_col},
                                    'pasteType': 'PASTE_FORMAT'}}]

    # Freeze top row and left colums
    freeze = [{'updateSheetProperties': {'properties': {'sheetId': s_id_bg,
                                                        'gridProperties': {'frozenRowCount': 1,
                                                                           'frozenColumnCount': header.index('Placement')+1}},
                                         'fields': 'gridProperties(frozenRowCount, frozenColumnCount)'}}]

    # Add filter
    basic_filter = [{'setBasicFilter': {'filter': {'range': {'sheetId': s_id_bg, 
                                                   'startRowIndex': 0,
                                                   'endRowIndex': n_row-1,
                                                   'startColumnIndex': 0, 
                                                   'endColumnIndex': n_col}}}}]

    # Add a filter view per cm
    i_cm = header.index('CM')
    cms = []
    for i in range(1, len(values)-1):
        cms.append(values[i][i_cm])
    cms = set(cms)

    filter_views = []
    for cm in cms:
        filter_view_name = cm.split()[0].lower()
        filter_view_dict = {'addFilterView': {'filter': {'title': filter_view_name, 
                                                         'range': {'sheetId': s_id_bg,
                                                                   'startRowIndex': 0,
                                                                   'endRowIndex': n_row-1,
                                                                   'startColumnIndex': 0,
                                                                   'endColumnIndex': n_col},
                                                         'criteria': {i_cm: {'condition': {'type': 'TEXT_EQ', 
                                                                                           'values': [{'userEnteredValue': cm}]}}}}}}
        filter_views.append(filter_view_dict)

    # Add a filter view for expedited invoice
    filter_view_dict = {'addFilterView': {'filter': {'title': 'exp inv', 
                                                     'range': {'sheetId': s_id_bg,
                                                               'startRowIndex': 0,
                                                               'endRowIndex': n_row-1,
                                                               'startColumnIndex': 0,
                                                               'endColumnIndex': n_col},
                                                     'criteria': {header.index('Expedited Invoice'): {'condition': {'type': 'TEXT_EQ', 
                                                                                                                    'values': [{'userEnteredValue': 'TRUE'}]}}}}}}
    filter_views.append(filter_view_dict)

    # Column width
    width = []
    for col in header:
        if col == 'OLI':
            size = 55
        elif col == 'Campaign Name':
            size = 140
        elif col == 'Line Item Number':
            size = 50
        elif col == 'Placement':
            size = 200
        else:
            size = 90

        width_dict = {'updateDimensionProperties': {'range': {'sheetId': s_id_bg,
                                                              'dimension': 'COLUMNS',
                                                              'startIndex': header.index(col),
                                                              'endIndex': header.index(col)+1},
                                                    'properties': {'pixelSize': size},
                                                    'fields': 'pixelSize'}}
        width.append(width_dict)

    # Send requests
    all_requests = header_format + content_format + footer_format
    all_requests += freeze + basic_filter + filter_views + width

    result = service.spreadsheets().batchUpdate(spreadsheetId=file_id, body={'requests': all_requests}).execute()

    # Delete formatting template sheet
    result = service.spreadsheets().batchUpdate(spreadsheetId=file_id, body={'requests': [{'deleteSheet': {'sheetId': s_id_template}}]}).execute()

    return

def add_cpuv_me2bg_gsheet(year, mo, bg_ss_id):

    ##############################################################
    # Make data
    ##############################################################

    # Get CPUV Goals Sheet
    DIR_PICKLES = PATH2PICKLES + '/' + PREFIX4ALWAYS_UP2DATE + str(year) + '_' + str(mo).zfill(2)
    with open(DIR_PICKLES + '/' + 'cpuv_goals.pickle', 'rb') as f:
        cpuv_goals = pickle.load(f)

    # Columns from Goals Sheet
    site_goal_col_list = []
    for col in cpuv_goals.columns.tolist():
        if ('Goal' in col) & (col not in ['Goal', 'Goal Check', 'Max at Goal']):
            site_goal_col_list.append(col)

    me = cpuv_goals[['OLI', 'BBR', 'Campaign Name', 'Line Description', 'Price Calculation Type', '1st Line Item', 'Goal'] + site_goal_col_list]

    # Columns to add
    site_delivery_col_list = [col.replace(' Goal', ' Delivery') for col in site_goal_col_list]

    for col in ['Hit the Goal?', 'Delivery', 'Delivery (Paid + AV)', 'Goal (Paid + AV)'] + site_delivery_col_list:
        me[col] = ''

    # Clean up
    for col in site_goal_col_list:
        me.loc[me[col] == 0, col] = ''

    header = ['OLI', 'BBR', 'Campaign Name', 'Line Description', 'Price Calculation Type', 'Hit the Goal?', '1st Line Item',
              'Delivery', 'Delivery (Paid + AV)', 'Goal (Paid + AV)', 'Goal']

    for i in range(len(site_goal_col_list)):
        header.append(site_goal_col_list[i])
        header.append(site_delivery_col_list[i])

    me = me[header]

    # Add formulas
    def col_name2col_letter(col_name):
        col_num = header.index(col_name) + 1
        return opx.utils.get_column_letter(col_num)

    col_delivery = col_name2col_letter('Delivery')
    col_goal = col_name2col_letter('Goal')
    col_1st_line = col_name2col_letter('1st Line Item')

    def get_hit_goal(i):
        row = str(i + 2)
        return '=IF(' + col_delivery + row + '>=' + col_goal + row + ', "Y", "N")'

    site_delivery_col_letter_list = [col_name2col_letter(col) for col in site_delivery_col_list]
    def get_delivery_paid_av(i):
        row = str(i + 2)
        cell_list = [(j + row) for j in site_delivery_col_letter_list]
        return '=' + '+'.join(cell_list)

    def get_goal_paid_av(i):
        row = str(i + 2)
        next_row = str(i + 3)
        output = '=IF(AND(' + col_1st_line + row + '=1, ' + col_1st_line + next_row + '=0), '
        output += col_goal + row + '+' + col_goal + next_row + ', '
        output += 'IF(' + col_1st_line + row + '=0, 0, ' + col_goal + row + '))'
        return output

    me['Hit the Goal?'] = [get_hit_goal(i) for i in range(len(me))]
    me['Delivery (Paid + AV)'] = [get_delivery_paid_av(i) for i in range(len(me))]
    me['Goal (Paid + AV)'] = [get_goal_paid_av(i) for i in range(len(me))]

    ##############################################################
    # Upload values to BG Google Sheet as 'CPUV ME' sheet
    ##############################################################

    # Replace null with empty string. Otherwise upload to Google Sheet fails.
    me = me.fillna('')
    me_values = [me.columns.tolist()] + me.values.tolist()

    # Upload
    service = get_gsheet_service()

    sheet_name = 'CPUV ME'
    sheet_id = gsheet_create_sheet(sheet_name, bg_ss_id)
    result = service.spreadsheets().values().update(spreadsheetId=bg_ss_id, range=sheet_name,
                                                    valueInputOption='USER_ENTERED',
                                                    body={'values': me_values}).execute()

    ##############################################################
    # Format sheet
    ##############################################################

    n_row = len(me) + 1
    n_col = len(header)

    # Freeze header row and columns up to Line Description
    freeze = [{'updateSheetProperties': {'properties': {'sheetId': sheet_id,
                                                        'gridProperties': {'frozenRowCount': 1,
                                                                           'frozenColumnCount': header.index('Line Description') + 1}},
                                         'fields': 'gridProperties(frozenRowCount, frozenColumnCount)'}}]

    # Wrap header
    wrap = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'wrapStrategy': 'WRAP'}}] * n_col}],
                             'fields': 'userEnteredFormat.wrapStrategy',
                             'range': {'sheetId': sheet_id,
                                       'startRowIndex': 0,
                                       'endRowIndex': 1,
                                       'startColumnIndex': 0,
                                       'endColumnIndex': n_col}}}]

    # Color header
    yellow = {'red': 1, 'green': 1, 'blue': .729, 'alpha': 1}
    orange = {'red': 1, 'green': .875, 'blue': .729, 'alpha': 1}
    blue = {'red': .729, 'green': .882, 'blue': 1, 'alpha': 1}
    green = {'red': .729, 'green': 1, 'blue': .788, 'alpha': 1}
    purple = {'red': .867, 'green': .827, 'blue': .933, 'alpha': 1}
    pink = {'red': 1, 'green': .882, 'blue': .902, 'alpha': 1}
    light_blue = {'red': .827, 'green': .933, 'blue': .922, 'alpha': 1}

    site_color_dict = {0: orange, 1: blue, 2: green, 3: purple, 4: pink, 5: light_blue}

    def make_color_dict(start_col_name, n2repeat, rgba):
        return {'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'backgroundColor': rgba}}] * n2repeat}],
                                'fields': 'userEnteredFormat.backgroundColor',
                                'range': {'sheetId': sheet_id,
                                          'startRowIndex': 0,
                                          'endRowIndex': 1,
                                          'startColumnIndex': header.index(start_col_name),
                                          'endColumnIndex': header.index(start_col_name) + n2repeat}}}

    color = []
    color.append(make_color_dict(header[0], header.index('Goal')+1, yellow))
    for i in range(len(site_goal_col_list)):
        i_mod = i % len(site_color_dict)
        color.append(make_color_dict(site_goal_col_list[i], 2, site_color_dict[i_mod]))

    # Font
    font = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'textFormat': {'fontFamily': 'Roboto',
                                                                                        'fontSize': 9}}}] * n_col}] * n_row,
                             'fields': 'userEnteredFormat.textFormat(fontFamily, fontSize)',
                             'range': {'sheetId': sheet_id,
                                       'startRowIndex': 0,
                                       'endRowIndex': n_row,
                                       'startColumnIndex': 0,
                                       'endColumnIndex': n_col}}}]

    # Value formatting for UVs
    value_formatting = [{'repeatCell': {'range': {'sheetId': sheet_id,
                                                  'startRowIndex': 1,
                                                  'endRowIndex': n_row,
                                                  'startColumnIndex': header.index('Delivery'),
                                                  'endColumnIndex': n_col},
                                        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'NUMBER',
                                                                                        'pattern': '###,###,##0'}}},
                                        'fields': 'userEnteredFormat.numberFormat'}}]

    # Add filter
    basic_filter = [{'setBasicFilter': {'filter': {'range': {'sheetId': sheet_id,
                                                             'startRowIndex': 0,
                                                             'endRowIndex': n_row,
                                                             'startColumnIndex': 0,
                                                             'endColumnIndex': n_col}}}}]

    # Column width
    width = [{'updateDimensionProperties': {'range': {'sheetId': sheet_id,
                                                      'dimension': 'COLUMNS',
                                                      'startIndex': header.index('Delivery'),
                                                      'endIndex': n_col},
                                            'properties': {'pixelSize': 75},
                                            'fields': 'pixelSize'}}]

    # Send all requests
    requests = freeze + wrap + color + font + value_formatting + basic_filter + width
    result = service.spreadsheets().batchUpdate(spreadsheetId=bg_ss_id, body={'requests': requests}).execute()

    return

def update_bg_cpuv_formulas(file_id):

    service = get_gsheet_service()
    sheet_name = 'BG'

    # Make a request to Google Sheet
    result = service.spreadsheets().values().get(
        spreadsheetId=file_id, range=sheet_name, valueRenderOption='UNFORMATTED_VALUE').execute()
    values = result.get('values', [])

    # If CPUV, update formulas
    header = values[0]
    i_type = header.index('Rate Type')

    def col_name2col_letter(col_name):
        col_num = header.index(col_name) + 1
        return opx.utils.get_column_letter(col_num)

    col_letter_fpu = col_name2col_letter('First Party Units')
    col_letter_oli = col_name2col_letter('OLI')
    col_letter_tpu = col_name2col_letter('Third Party Impressions')

    upload_data = []
    for i in range(len(values)):
        row = values[i]
        if row[i_type] != 'CPUV':
            continue

        i_row = str(i + 1)
        
        # First Party Units
        cell = sheet_name + '!' + col_letter_fpu + i_row
        value = "=IFERROR(VLOOKUP(" + col_letter_oli + i_row + ",'CPUV ME'!$A:$H,8,false), 0)"
        upload_data.append({'range': cell, 'majorDimension': 'ROWS', 'values': [[value]]})

        # Third Party Units
        cell = sheet_name + '!' + col_letter_tpu + i_row
        value = "=" + col_letter_fpu + i_row
        upload_data.append({'range': cell, 'majorDimension': 'ROWS', 'values': [[value]]})

    result = service.spreadsheets().values().batchUpdate(spreadsheetId=file_id, body={'valueInputOption': 'USER_ENTERED', 'data': upload_data}).execute()
    
    return

def get_bg(year, mo):

    ##############################################################
    folder_id = '0B71ox_2Qc7gmWkNWM21YUU9MSEU'
    ##############################################################

    months_dict = {1: 'January', 2: 'Febuary', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
                   7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}

    ss_name = months_dict[mo] + ' ' + str(year) + ' Billing Grid'
    ss_id = gdrive_get_file_id_by_name(ss_name, folder_id)

    # Make a request to Google Sheet
    service = get_gsheet_service()
    result = service.spreadsheets().values().get(
        spreadsheetId=ss_id, range='BG', valueRenderOption='UNFORMATTED_VALUE').execute()
    values = result.get('values', [])

    df = pd.DataFrame(values[1: len(values) - 1], columns=values[0])  # Exclude the last row (Grand Total)
    return df   

def adjust_unfinalized_bg(bg):
    df = bg

    def fill_in_temp_billed(row):
        system = row['Third Party System']
        third = row['Third Party Impressions']
        billed = row['Billed Units']
        booked = row['Booked Impressions']

        if system == 'DFP':
            return billed
        if (third is None) or (third == ''):
            return int(booked * 0.5)
        return billed

    df['Billed Units'] = df.apply(lambda row: fill_in_temp_billed(row), axis=1)
    
    return df

##############################################################
# Ask.com TP
##############################################################

def make_ask_tp_as_excel(year, mo):
   
    ############################################################## 
    # Prep
    ##############################################################

    output_file_name = 'for_ask_tp_' + str(year) + '_' + str(mo).zfill(2) + '.xlsx'
    
    ##############################################################
    # Get data
    ##############################################################

    # Labeled dfp report
    row_dfp = run_dfp_mtd_ask_tp_query(year, mo)
    labeled_dfp = label_dfp_mtd_all(row_dfp)

    dummy_cpuv = pd.DataFrame(columns=['BBR', 'Base Rate', 'Campaign Name', 'Date', 'Line Description', 'Original Report Tab Name', 
                                       'Report Tab Name', 'Site', 'UV entered', 'UVs'])
    all1 = make_all1(labeled_dfp, dummy_cpuv, None)

    df = all1[pd.notnull(all1['Ad unit'])]  # Only include DFP data, no UV data

    # Extract Ask TP imps
    df = df[df['Site'] == 'HL']

    df.loc[(df['Ad unit'].str.contains('ask', case=False) & df['Ad unit'].str.contains('tp', case=False)), 'Ask TP Domain'] = 'Ask.com'
    df.loc[df['Ad unit'].str.contains('.ask'), 'Ask TP Domain'] = 'Ask.com'
    df.loc[df['Ad unit'].str.contains('.sas'), 'Ask TP Domain'] = 'Search.Ask.com'
    df.loc[df['Ad unit'].str.contains('.mwy'), 'Ask TP Domain'] = 'MyWay.com'
    
    df = df[pd.notnull(df['Ask TP Domain'])]

    ##############################################################
    # Main sheet
    ##############################################################

    # Format
    groupby_col = ['Ask TP Domain', 'Ad unit', 'Price Calculation Type', 'Advertiser', 'Order', 'Line item', 'Base Rate']
    value_col = ['Impressions/UVs']
    rename_dict = {'Base Rate': 'Rate', 'Impressions/UVs': 'Impressions'}

    for col in groupby_col:
        df.loc[pd.isnull(df[col]), col] = 'N/A'

    df = df[groupby_col + value_col].groupby(groupby_col).sum().reset_index().rename(columns=rename_dict)

    # Add Revenue Type
    def add_rev_type(row):
        if row['Price Calculation Type'] == 'CPM':
            return 'DAS'
        if row['Advertiser'] in ['Programmatic Partners', 'OpenX']:
            return 'Programmatic'
        if row['Advertiser'].startswith('Prebid'):
            return 'Prebid'
        return ''

    df['Revenue Type'] = df.apply(lambda row: add_rev_type(row), axis=1)

    # Add Gross Revenue, to be populated with excel formula
    df['Gross Revenue'] = ''

    # Clean up
    header = ['Ask TP Domain', 'Ad unit', 'Advertiser', 'Order', 'Line item', 'Revenue Type', 'Rate', 'Impressions', 'Gross Revenue']
    sortby = ['Ask TP Domain', 'Advertiser', 'Order', 'Line item', 'Rate']
    df = df[header].sort_values(sortby).reset_index(drop=True)

    # Add formula for Rate
    order_col_letter = opx.utils.get_column_letter(header.index('Order') + 1)

    def update_rate(row):
        if row['Revenue Type'] == 'DAS':
            return row['Rate']
        if row['Revenue Type'] == 'Programmatic':
            str_row_num = str(row.name + 2)
            return "=VLOOKUP(" + order_col_letter + str_row_num + ",'Programmatic eCPMs'!$A:$B,2,FALSE)"
        if row['Revenue Type'] == 'Prebid':
            return "='Prebid eCPMs'!B1"
        return 0

    df['Rate'] = df.apply(lambda row: update_rate(row), axis=1)

    # Add formula for Revenue
    imps_col_letter = opx.utils.get_column_letter(header.index('Impressions') + 1)
    rate_col_letter = opx.utils.get_column_letter(header.index('Rate') + 1)

    df['Gross Revenue'] = ['=' + imps_col_letter + str(i+2) + '/1000*' + rate_col_letter + str(i+2) for i in range(len(df))]

    ##############################################################
    # eCPM sheets
    ##############################################################

    # Sheet for Programmatic eCPMs (To be filled in by Chris later, average eCPM per DFP Order)
    df_prog = df[df['Revenue Type'] == 'Programmatic'][['Order']].drop_duplicates().sort_values(['Order'])
    df_prog['eCPM'] = None

    # Sheet for Prebid eCPMs (To be filled in by Chris later, average eCPM of all)
    df_prebid = pd.DataFrame(columns=['Average'])

    ##############################################################
    # Write to excel
    ##############################################################

    writer = pd.ExcelWriter(output_file_name)
    df.to_excel(writer, 'main', index=False)
    df_prog.to_excel(writer, 'Programmatic eCPMs', index=False)
    df_prebid.to_excel(writer, 'Prebid eCPMs', index=False)
    writer.save()

    return output_file_name

def upload_ask_tp2gdrive(year, mo, excel_file_name):
    
    ##############################################################
    folder_id = '0B71ox_2Qc7gmSFBTUjctXzJMRFU'
    ##############################################################

    gsheet_file_name = 'Ask.com_TP_' + str(year) + '_' +  str(mo).zfill(2)
    
    file_id = save_excel_as_gsheet_in_gdrive(gsheet_file_name, folder_id, excel_file_name)

    return file_id

def format_ask_tp(file_id):

    service = get_gsheet_service()

    ##############################################################
    # Make a dictionary of basic sheet info
    ##############################################################

    sheet_dict = {}

    ss_metadata = service.spreadsheets().get(spreadsheetId=file_id).execute()
    for s_metadata in ss_metadata['sheets']:
        # Sheet name and id
        sheet_name = s_metadata['properties']['title']
        sheet_id = s_metadata['properties']['sheetId'] 

        # Get values > header, # of col, # of row
        result = service.spreadsheets().values().get(
            spreadsheetId=file_id, range=sheet_name, valueRenderOption='UNFORMATTED_VALUE').execute()
        values = result.get('values', [])

        header = values[0]
        n_col = len(header)
        n_row = len(values)

        # Store as a tuple
        sheet_dict[sheet_name] = (sheet_id, header, n_col, n_row)        

    ##############################################################
    # Helper functions etc.
    ##############################################################

    def make_vf_dict(sheet_name, col, format_type):
        sheet_id, header, n_col, n_row = sheet_dict[sheet_name]
        i_col = header.index(col)

        if format_type == 'integer':
            pattern = '###,###,##0'
        elif format_type == 'dollar':
            pattern = '$#,###,##0.00'
        
        return {'repeatCell': {'range': {'sheetId': sheet_id,
                                         'startRowIndex': 1,
                                         'endRowIndex': n_row,
                                         'startColumnIndex': i_col,
                                         'endColumnIndex': i_col+1},
                               'cell': {'userEnteredFormat': {'numberFormat': {'type': 'NUMBER',
                                                                               'pattern': pattern}}},
                               'fields': 'userEnteredFormat.numberFormat'}}
    
    # color dicts
    green = {'red': .851, 'green': .918, 'blue': .827, 'alpha': 1}
    yellow = {'red': 1, 'green': .949, 'blue': .8, 'alpha': 1}

    def make_color_header_list(sheet_name):
        sheet_id, header, n_col, n_row = sheet_dict[sheet_name]

        return [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'backgroundColor': green}}] * n_col}],
                                 'fields': 'userEnteredFormat.backgroundColor',
                                 'range': {'sheetId': sheet_id,
                                           'startRowIndex': 0,
                                           'endRowIndex': 1,
                                           'startColumnIndex': 0,
                                           'endColumnIndex': n_col}}}]

    ##############################################################
    # Format main sheet
    ##############################################################

    sheet_name = 'main'
    sheet_id, header, n_col, n_row = sheet_dict[sheet_name]

    # Freeze top row
    freeze = [{'updateSheetProperties': {'properties': {'sheetId': sheet_id,
                                                        'gridProperties': {'frozenRowCount': 1}},
                                         'fields': 'gridProperties(frozenRowCount)'}}]

    # Add filter
    basic_filter = [{'setBasicFilter': {'filter': {'range': {'sheetId': sheet_id,
                                                   'startRowIndex': 0,
                                                   'endRowIndex': n_row,
                                                   'startColumnIndex': 0,
                                                   'endColumnIndex': n_col}}}}]

    # Wrap header
    wrap = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'wrapStrategy': 'WRAP'}}] * n_col}],
                             'fields': 'userEnteredFormat.wrapStrategy',
                             'range': {'sheetId': sheet_id,
                                       'startRowIndex': 0,
                                       'endRowIndex': 1,
                                       'startColumnIndex': 0,
                                       'endColumnIndex': n_col}}}]
    
    # Color header
    color = make_color_header_list(sheet_name)

    # Value formatting
    value_formatting = []
    value_formatting.append(make_vf_dict(sheet_name, 'Rate', 'dollar'))
    value_formatting.append(make_vf_dict(sheet_name, 'Impressions', 'integer'))
    value_formatting.append(make_vf_dict(sheet_name, 'Gross Revenue', 'dollar'))

    # Width
    width = [{'updateDimensionProperties': {'range': {'sheetId': sheet_id,
                                                      'dimension': 'COLUMNS',
                                                      'startIndex': 0,
                                                      'endIndex': n_col},
                                            'properties': {'pixelSize': 90},
                                            'fields': 'pixelSize'}}]
   
    # Clip
    clip = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'wrapStrategy': 'CLIP'}}] * n_col}] * (n_row-1),
                             'fields': 'userEnteredFormat.wrapStrategy',
                             'range': {'sheetId': sheet_id,
                                       'startRowIndex': 1,
                                       'endRowIndex': n_row,
                                       'startColumnIndex': 0,
                                       'endColumnIndex': n_col}}}]
 
    # Send requests
    requests = freeze + basic_filter + wrap + color + value_formatting + width + clip
    result = service.spreadsheets().batchUpdate(spreadsheetId=file_id, body={'requests': requests}).execute()

    ##############################################################
    # Format programmatic sheet
    ##############################################################

    sheet_name = 'Programmatic eCPMs'
    sheet_id, header, n_col, n_row = sheet_dict[sheet_name]

    # Color header
    color_header = make_color_header_list(sheet_name)

    color_rates = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'backgroundColor': yellow}}]}] * (n_row-1),
                                   'fields': 'userEnteredFormat.backgroundColor',
                                   'range': {'sheetId': sheet_id,
                                             'startRowIndex': 1,
                                             'endRowIndex': n_row,
                                             'startColumnIndex': 1,
                                             'endColumnIndex': 2}}}]

    # Value formatting
    value_formatting = []
    value_formatting.append(make_vf_dict(sheet_name, 'eCPM', 'dollar'))

    # Width
    width = [{'updateDimensionProperties': {'range': {'sheetId': sheet_id,
                                                      'dimension': 'COLUMNS',
                                                      'startIndex': 0,
                                                      'endIndex': 1},
                                            'properties': {'pixelSize': 240},
                                            'fields': 'pixelSize'}}]

    # Send requests
    requests = color_header + color_rates + value_formatting + width
    result = service.spreadsheets().batchUpdate(spreadsheetId=file_id, body={'requests': requests}).execute()

    ##############################################################
    # Format prebid sheet
    ##############################################################

    sheet_name = 'Prebid eCPMs'
    sheet_id, header, n_col, n_row = sheet_dict[sheet_name]

    # Color header, just cell A1
    # Color cell for Chris, just cell B1
    color = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'backgroundColor': green}}, 
                                                   {'userEnteredFormat': {'backgroundColor': yellow}}]}],
                              'fields': 'userEnteredFormat.backgroundColor',
                              'range': {'sheetId': sheet_id,
                                        'startRowIndex': 0,
                                        'endRowIndex': 1,
                                        'startColumnIndex': 0,
                                        'endColumnIndex': 2}}}]

    # Value formatting
    value_formatting = [{'repeatCell': {'range': {'sheetId': sheet_id,
                                        'startRowIndex': 0,
                                        'endRowIndex': 1,
                                        'startColumnIndex': 1,
                                        'endColumnIndex': 2},
                         'cell': {'userEnteredFormat': {'numberFormat': {'type': 'NUMBER',
                                                                                 'pattern': '$#,###,##0.00'}}},
                         'fields': 'userEnteredFormat.numberFormat'}}]

    # Send requests
    requests = color + value_formatting 
    result = service.spreadsheets().batchUpdate(spreadsheetId=file_id, body={'requests': requests}).execute()

    ##############################################################
    # Create summary sheet (piv table of main sheet)
    # Reference: https://developers.google.com/sheets/api/samples/pivot-tables
    ##############################################################
    
    # Source sheet info
    sheet_id, header, n_col, n_row = sheet_dict['main']

    # Create summary sheet and move to far left
    piv_sheet_id =  gsheet_create_sheet('summary', file_id)
    move = [{'updateSheetProperties': {'properties': {'sheetId': piv_sheet_id, 
                                                      'index': 0}, 
                                       'fields': 'index'}}]

    # Insert pivot table
    pivt = [{'updateCells': {'rows': {'values': [{'pivotTable': {'source': {'sheetId': sheet_id,
                                                                            'startRowIndex': 0,
                                                                            'endRowIndex': n_row,
                                                                            'startColumnIndex': 0,
                                                                            'endColumnIndex': n_col},
                                                                 'rows': [{'sourceColumnOffset': header.index('Ask TP Domain'),
                                                                           'showTotals': True,
                                                                           'sortOrder': 'ASCENDING',
                                                                           'valueBucket': {'buckets': [{'stringValue': 'Ask TP Domain'}]}},
                                                                          {'sourceColumnOffset': header.index('Revenue Type'),
                                                                           'showTotals': True,
                                                                           'sortOrder': 'ASCENDING',
                                                                           'valueBucket': {'buckets': [{'stringValue': 'Revenue Type'}]}}],
                                                                 'values': [{'summarizeFunction': 'SUM',
                                                                             'sourceColumnOffset': header.index('Impressions')},
                                                                            {'summarizeFunction': 'SUM',
                                                                             'sourceColumnOffset': header.index('Gross Revenue')}],
                                                                 'valueLayout': 'HORIZONTAL'}}]},
                             'start': {'sheetId': piv_sheet_id,
                                       'rowIndex': 0,
                                       'columnIndex': 0},
                             'fields': 'pivotTable'}}]

    requests = move + pivt
    result = service.spreadsheets().batchUpdate(spreadsheetId=file_id, body={'requests': requests}).execute()

##############################################################
# Site Report
##############################################################

def make_site_report_as_excel(year, mo, prefix4output):

    ##############################################################
    # Output file name
    ##############################################################

    today_date = datetime.now().date()
    today_date_str = str(today_date.month).zfill(2) + str(today_date.day).zfill(2) + str(today_date.year)

    months_dict = {1: 'January', 2: 'Febuary', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
                   7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}

    output_file_name = prefix4output + '_' + months_dict[mo] + '_' + str(year) + '_Site_Report_' + today_date_str + '.xlsx'

    ##############################################################
    # Get data
    ##############################################################

    DIR_PICKLES = PATH2PICKLES + '/' + PREFIX4ALWAYS_UP2DATE + str(year) + '_' + str(mo).zfill(2)

    with open(DIR_PICKLES + '/' + 'all1.pickle', 'rb') as f:
        all1 = pickle.load(f)

    with open(DIR_PICKLES + '/' + 'site_goals.pickle', 'rb') as f:
        site_goals = pickle.load(f)

    bg = get_bg(year, mo)
    bg = adjust_unfinalized_bg(bg)

    ##############################################################
    # Prep
    ##############################################################

    bbd_list = ['BBR', 'Brand', 'DAS Line Item Name']

    ##############################################################
    # Get adjusted (estimated 3rd party) impressions per site
    ##############################################################

    # Extract CPM and CPUV data from the 'all1' file. Aggregate the data.
    data = all1[(all1['Price Calculation Type'] == 'CPM') | (all1['Price Calculation Type'] == 'CPUV')]
    data.loc[pd.isnull(data['(DAS)BBR #']), '(DAS)BBR #'] = data.loc[pd.isnull(data['(DAS)BBR #']), '(Order)BBR #']
    data = data[['(DAS)BBR #', 'Brand', 'DAS Line Item Name', 'Site', 'Price Calculation Type', 'Impressions/UVs', 'Clicks']]
    data = data.rename(columns={'(DAS)BBR #': 'BBR', 'Impressions/UVs': 'Delivered'})
    data = data.groupby(bbd_list+['Site', 'Price Calculation Type']).sum().reset_index()

    # Add Discrepancy from BG
    bg_discrepancy = bg[['Advertiser', 'Placement', 'Discrepancy', 'BBR']]
    bg_discrepancy = bg_discrepancy.rename(columns={'Advertiser': 'Brand', 'Placement': 'DAS Line Item Name'})
    data = pd.merge(data, bg_discrepancy, how='left', on=bbd_list)

    # If Discrepancy is less than or equal to 5%, partners get paid for all they delivered
    # If Discrepancy is greater than 5%, partners get paid for all minus the discrepancy
    # For DSP, let 'Adjusted w/Discrepancy' = 'Delivered' regardless of Discrepancy
    data.loc[[isinstance(d, str) for d in data['Discrepancy']], 'Discrepancy'] = 0
    data.loc[data['Discrepancy'] == 1, 'Discrepancy'] = 0

    data['Adjusted w/ Discrepancy'] = data['Delivered'] * (1 - data['Discrepancy'])
    data.loc[data['Discrepancy'] <= 0.05, 'Adjusted w/ Discrepancy'] = data.loc[data['Discrepancy'] <= 0.05, 'Delivered']
    data.loc[pd.isnull(data['Discrepancy']), 'Adjusted w/ Discrepancy'] = data.loc[pd.isnull(data['Discrepancy']), 'Delivered']
    data['Adjusted w/ Discrepancy'] = [round(a_w_d, 0) for a_w_d in data['Adjusted w/ Discrepancy']]

    ##############################################################
    # Add site goals
    ##############################################################

    data = pd.merge(data, site_goals[bbd_list + ['Site', 'Site Goal']],
                    how='left', on=bbd_list+['Site'])
    data.loc[pd.isnull(data['Site Goal']), 'Site Goal'] = 0

    ##############################################################
    # HL & MNT delivery
    ##############################################################

    mnt = data[data['Site'] == 'Medical News Today']
    mnt = mnt[bbd_list + ['Adjusted w/ Discrepancy']]
    mnt = mnt.rename(columns={'Adjusted w/ Discrepancy': 'MNT Delivery'})

    hl = data[data['Site'] == 'HL']
    hl = hl[bbd_list + ['Adjusted w/ Discrepancy']]
    hl = hl.rename(columns={'Adjusted w/ Discrepancy': 'HL Delivery'})

    ##############################################################
    # HW delivery & capped at site goal
    # Cap EXCEPT FOR Drugs.com CPUV Microsite
    ##############################################################

    hw_sites = data[(data['Site'] != 'HL') & (data['Site'] != 'Medical News Today')]

    def cap_hw_delivery(row):
        if (row['Site'] == 'Drugs.com') & (row['Price Calculation Type'] == 'CPUV') & \
                ('Competitive Conquesting' not in row['DAS Line Item Name']):
            return row['Adjusted w/ Discrepancy']
        if row['Adjusted w/ Discrepancy'] > row['Site Goal']:
            return row['Site Goal']
        else:
            return row['Adjusted w/ Discrepancy']

    hw_sites['Capped'] = hw_sites.apply(lambda row: cap_hw_delivery(row), axis=1)

    ##############################################################
    # Calculation dataframe
    ##############################################################

    # Make
    calc_df = bg[['Advertiser', 'Placement', 'Billed Units', 'BBR']]
    calc_df['Billed Units'] = [round(bi, 0) for bi in calc_df['Billed Units']]
    calc_df = calc_df.rename(columns={'Advertiser': 'Brand',
                                      'Placement': 'DAS Line Item Name',
                                      'Billed Units': 'Total Billed'})

    # Add # of HW sites
    hw_count = data[(data['Site'] != 'HL') & (data['Site'] != 'Medical News Today') & (data['Adjusted w/ Discrepancy'] > 0)]
    hw_count = hw_count[bbd_list].groupby(bbd_list).size()
    hw_count = hw_count.reset_index().rename(columns={0: 'HW Count'})

    calc_df = pd.merge(calc_df, hw_count, how='left', on=bbd_list)
    calc_df.loc[pd.isnull(calc_df['HW Count']), 'HW Count'] = 0

    # Add HW delivery & capped at site goal
    hw_total = hw_sites[bbd_list + ['Adjusted w/ Discrepancy', 'Capped']]
    hw_total = hw_total.groupby(bbd_list).sum().reset_index()
    hw_total = hw_total.rename(columns={'Adjusted w/ Discrepancy': 'HW Delivery', 'Capped': 'HW Capped'})

    calc_df = pd.merge(calc_df, hw_total, how='left', on=bbd_list)
    calc_df.loc[pd.isnull(calc_df['HW Delivery']), 'HW Delivery'] = 0
    calc_df.loc[pd.isnull(calc_df['HW Capped']), 'HW Capped'] = 0

    # Add MNT delivery & capped at overall goal
    calc_df = pd.merge(calc_df, mnt, how='left', on=bbd_list)
    calc_df.loc[pd.isnull(calc_df['MNT Delivery']), 'MNT Delivery'] = 0

    calc_df['Total Billed - HW Capped'] = calc_df['Total Billed'] - calc_df['HW Capped']

    def fill_mnt_billed(row, hw_col):
        if row['MNT Delivery'] <= 0:
            return 0
        if row['Total Billed - ' + hw_col] < 0:
            return 0
        if row['MNT Delivery'] < row['Total Billed - ' + hw_col]:
            return row['MNT Delivery']
        return row['Total Billed - ' + hw_col]

    calc_df['MNT Billed'] = calc_df.apply(lambda row: fill_mnt_billed(row, 'HW Capped'), axis=1)

    # Add HL delivery & left billable if HW delivery is capped at site goal
    calc_df = pd.merge(calc_df, hl, how='left', on=bbd_list)
    calc_df.loc[pd.isnull(calc_df['HL Delivery']), 'HL Delivery'] = 0

    calc_df['HL Left Billable'] = calc_df['Total Billed'] - calc_df['HW Capped'] - calc_df['MNT Billed']

    # If site-specific, adjust
    # If HL didn't deliver left billable, adjust
    calc_df['HL Delivery - HL Left Billable'] = calc_df['HL Delivery'] - calc_df['HL Left Billable']

    def fill_hw_billed(row):
        if row['HW Count'] == 0:
            return 0
        if (row['HW Count'] == 1) & (row['MNT Delivery'] <= 0) & (row['HL Delivery'] <= 0):
            return row['Total Billed']
        if row['HL Delivery - HL Left Billable'] >= 0:
            return row['HW Capped']

        mnt_left_billable = row['MNT Delivery'] - row['MNT Billed']
        give2hw = -1 * row['HL Delivery - HL Left Billable'] - mnt_left_billable
        return row['HW Capped'] + give2hw

    calc_df['HW Billed'] = calc_df.apply(lambda row: fill_hw_billed(row), axis=1)

    # Recalculate MNT billed. Final HL billed.
    calc_df['Total Billed - HW Billed'] = calc_df['Total Billed'] - calc_df['HW Billed']
    calc_df['MNT Billed'] = calc_df.apply(lambda row: fill_mnt_billed(row, 'HW Billed'), axis=1)
    calc_df['HL Billed'] = calc_df['Total Billed'] - calc_df['HW Billed'] - calc_df['MNT Billed']

    # Check 1: This should be zero for all rows
    calc_df['Total Billed - (HW Billed + MNT Billed + HL Billed)'] = calc_df['Total Billed'] - calc_df['HW Billed'] - calc_df['MNT Billed'] - calc_df['HL Billed']
    # Check 2: If this is negative, look into it
    calc_df['HL Delivery - HL Billed'] = calc_df['HL Delivery'] - calc_df['HL Billed']

    # Clean up
    calc_df = calc_df[['BBR', 'Brand', 'DAS Line Item Name',
                       'Total Billed', 'Total Billed - (HW Billed + MNT Billed + HL Billed)',
                       'HL Delivery - HL Billed', 'HL Delivery - HL Left Billable',
                       'HL Delivery', 'HL Billed',
                       'HW Count', 'HW Delivery', 'HW Capped', 'HW Billed',
                       'MNT Delivery', 'MNT Billed']]

    # Output
    calc_df.to_csv('billing_calculation_' + str(year) + '_' + str(mo).zfill(2) + '.csv', index=False)

    ##############################################################
    # Add billed to data
    ##############################################################

    # Round 1: Straightforward
    def fill_in_billed(row):
        mini_calc_df = calc_df[(calc_df['BBR'] == row['BBR']) &
                               (calc_df['Brand'] == row['Brand']) &
                               (calc_df['DAS Line Item Name'] == row['DAS Line Item Name'])]

        if len(mini_calc_df) == 0:
            return 'Not in BG'

        def value_of(col_in_mini_calc_df):
            return mini_calc_df[col_in_mini_calc_df].values[0]

        site = row['Site']

        if site == 'HL':
            return value_of('HL Billed')
        if site == 'Medical News Today':
            return value_of('MNT Billed')
        if (site == 'Drugs.com') & (row['Price Calculation Type'] == 'CPUV') & ('Competitive Conquesting' not in row['DAS Line Item Name']):
            return row['Adjusted w/ Discrepancy']
        if value_of('HW Count') == 1:
            return value_of('HW Billed')
        if value_of('HW Capped') == value_of('HW Billed'):
            return min(row['Adjusted w/ Discrepancy'], row['Site Goal'])
        return 'Check'

    data['Billed'] = data.apply(lambda row: fill_in_billed(row), axis=1)

    # Round 2: If 2 or more HW sites need to be adjusted
    to_check = data[data['Billed'] == 'Check'].drop_duplicates(bbd_list)
    for i in range(len(to_check)):
        mini_to_check = to_check.iloc[i]
        mini_data = data[(data['BBR'] == mini_to_check['BBR']) &
                         (data['Brand'] == mini_to_check['Brand']) &
                         (data['DAS Line Item Name'] == mini_to_check['DAS Line Item Name']) &
                         (data['Site'] != 'HL') &
                         (data['Site'] != 'Medical News Today')]

        # If site didn't over-deliver, let billed = delivery
        mini_data['UD'] = mini_data['Site Goal'] - mini_data['Adjusted w/ Discrepancy']
        sum_non_od_hw_billed = 0
        for j in range(len(mini_data)):
            per_site = mini_data.iloc[j]
            delivery = per_site['Adjusted w/ Discrepancy']
            if per_site['UD'] >= 0:
                data.loc[(data['BBR'] == mini_to_check['BBR']) &
                         (data['Brand'] == mini_to_check['Brand']) &
                         (data['DAS Line Item Name'] == mini_to_check['DAS Line Item Name']) &
                         (data['Site'] == per_site['Site']), 'Billed'] = delivery
                sum_non_od_hw_billed += delivery

        # If only 1 site over-delivered, let billed = leftover
        if len(mini_data[mini_data['UD'] < 0]) == 1:
            site = mini_data[mini_data['UD'] < 0]['Site'].values[0]

            mini_calc_df = calc_df[(calc_df['BBR'] == mini_to_check['BBR']) &
                                   (calc_df['Brand'] == mini_to_check['Brand']) &
                                   (calc_df['DAS Line Item Name'] == mini_to_check['DAS Line Item Name'])]
            hw_billed = mini_calc_df['HW Billed'].values[0]

            hw_od_billed = hw_billed - sum_non_od_hw_billed
            data.loc[(data['BBR'] == mini_to_check['BBR']) &
                     (data['Brand'] == mini_to_check['Brand']) &
                     (data['DAS Line Item Name'] == mini_to_check['DAS Line Item Name']) &
                     (data['Site'] == site), 'Billed'] = hw_od_billed

    ##############################################################
    # Add HL rows not in data
    # Most are to be changed to site = DSP (WWW lines)
    # Others are non-CPM/CPUV lines, like Flat-fee
    ##############################################################

    hl_in_data = hl[bbd_list]
    hl_in_data['HL delivery in all1'] = 1

    calc_df = pd.merge(calc_df, hl_in_data, how='left', on=bbd_list)
    calc_df.loc[pd.isnull(calc_df['HL delivery in all1']), 'HL delivery in all1'] = 0

    hl2add = calc_df[(calc_df['HL delivery in all1'] == 0) & (calc_df['HL Billed'] > 0)]
    hl2add = hl2add[bbd_list + ['HL Billed']].rename(columns={'HL Billed': 'Billed'})
    hl2add['Site'] = 'HL'
    data = pd.concat([data, hl2add])

    ##############################################################
    # Refresh Site Goal column and add RevShare column
    # Temporarily store Non-standard Site Rate in Net Site Expense column
    ##############################################################

    site_goals = site_goals[bbd_list + ['Site', 'Site Goal', 'Non-standard Site Rate']]
    data = data.drop('Site Goal', axis=1)
    data = pd.merge(data, site_goals, how='left', on=bbd_list+['Site'])
    data = data.rename(columns={'Non-standard Site Rate': 'Net Site Expense'})

    revshare_dict = get_revshare_dict()
    revshare_dict['HL'] = 0  # Need to revisit, to change the function
    revshare_dict['*Fill in*'] = '*Fill in*'
    data['RevShare'] = [revshare_dict[site] for site in data['Site']]

    ##############################################################
    # Add DAS columns
    ##############################################################

    das_month = str(mo) + '/' + str(year)
    das = make_das(use_scheduled_units=True, export=False)

    das_1 = das[das[das_month] > 0][['BBR', 'Campaign Name', 'Flight Type', 'Brand', 'Account Name', 'Agency', 'IO Number', 'Start Date', 'End Date',
                                     'Opportunity Owner', '2nd Opportunity Owner', 'Campaign Manager', 'Line Item Number', 'Line Description', 'Price Calculation Type',
                                     'Sales Price', 'Base Rate', 'Baked-In Production Rate', das_month, 'Customer Billing ID', 'Customer Billing Name']]
    das_1 = das_1.rename(columns={'Line Description': 'DAS Line Item Name', das_month: 'Total Goal'})

    data = data.drop('Price Calculation Type', axis=1)
    data = pd.merge(data, das_1, how='left', on=bbd_list)

    # Reformat the Sales Contact columns
    def sales_rep(row):
        rep1 = row['Opportunity Owner']
        rep2 = row['2nd Opportunity Owner']
        if isinstance(rep1, float):
            return rep1
        if rep2 == 'N/A':
            return rep1
        else:
            return rep1 + ', ' + rep2

    data['Sales Contact'] = data.apply(lambda row: sales_rep(row), axis=1)

    ##############################################################
    # For SEM campaigns, change Site from HL to *Fill in*
    ##############################################################

    data.loc[data['Campaign Manager'] == 'SEM', 'Site'] = '*Fill in*'

    ##############################################################
    # Add Total Billable from BG
    ##############################################################

    data = pd.merge(data, calc_df[bbd_list + ['Total Billed']], how='left', on=bbd_list)

    ##############################################################
    # Add Finalized?
    # If expedited invoice, 'Exp Inv'. Else 0.
    ##############################################################

    def is_expedited_invoice(ei):
        if isinstance(ei, bool):
            return ei
        if isinstance(ei, str):
            if ei.lower == 'true':
                return True
            return False
        return False

    exp_inv_bbr_list = bg[[is_expedited_invoice(ei) for ei in bg['Expedited Invoice']]]['BBR']
    exp_inv_bbr_list = list(set(exp_inv_bbr_list))

    def get_finalized(row):
        if row['BBR'] in exp_inv_bbr_list:
            return 'Exp Inv'
        return 0    

    data['Finalized?'] = data.apply(lambda row: get_finalized(row), axis=1)
 
    ##############################################################
    # Additional labeling
    ##############################################################

    data['HCP'] = ''
    data['HL/HW'] = 'HW'
    data['Own/Partner/DSP'] = ''

    # Label HCP
    data.loc[pd.notnull(data['Campaign Name']) & data['Campaign Name'].str.contains('HCP', case=False), 'HCP'] = 'HCP'

    # Add HL/HW
    data.loc[(data['Site'] == 'HL') | (data['Site'] == 'Medical News Today'), 'HL/HW'] = 'HL'

    # Add HL/Partner/DSP
    data.loc[(data['Site'] == 'HL') | (data['Site'] == 'Medical News Today'), 'Own/Partner/DSP'] = 'Own'
    data.loc[data['Campaign Manager'] == 'SEM', 'Own/Partner/DSP'] = 'DSP'
    data.loc[data['Own/Partner/DSP'] == '', 'Own/Partner/DSP'] = 'Partner'

    ##############################################################
    # Add columns to be calculated (Want formulas in Excel)
    ##############################################################

    data['Gross Site Revenue (Does Not Include Production Fee)'] = ''
    data['Production Fee'] = ''
    data['Gross Revenue (Includes Production Fee)'] = ''
    data['Billed > Delivered'] = ''

    ##############################################################
    # Rename columns and sort
    ##############################################################

    rename_dict = {'BBR': 'BBR #', 'Billed': 'Billed Impressions/UVs', 'Brand': 'Advertiser',
                   'Line Item Number': 'Line Item #', 'DAS Line Item Name': 'Line Item Name',
                   'Price Calculation Type': 'Unit', 'Baked-In Production Rate': 'Baked-in Production',
                   'Sales Price': 'Gross Rate', 'Start Date': 'Flight Start Date',
                   'End Date': 'Flight End Date', 'Flight Type': 'Goal Breakdown',
                   'Account Name': 'Parent', 'Total Billed': 'Total Billable'}

    header = ['Finalized?', 'BBR #', 'Advertiser', 'Campaign Name', 'Unit', 'Line Item #', 'Line Item Name', 'Site',
              'Billed Impressions/UVs', 'Site Goal', 'Adjusted w/ Discrepancy', 'Discrepancy', 'Delivered',
              'Goal Breakdown', 'Total Billable', 'Total Goal', 'Base Rate', 'Gross Site Revenue (Does Not Include Production Fee)',
              'RevShare', 'Net Site Expense', 'Baked-in Production', 'Production Fee', 'Gross Rate',
              'Gross Revenue (Includes Production Fee)', 'Own/Partner/DSP', 'HL/HW', 'HCP', 'Parent', 'Agency',
              'Customer Billing ID', 'Customer Billing Name', 'IO Number', 'Flight Start Date', 'Flight End Date',
              'Sales Contact', 'Billed > Delivered', 'Clicks']

    sortby = ['Advertiser', 'BBR #', 'Line Item #', 'Line Item Name', 'Site']

    data = data.rename(columns=rename_dict)[header].sort_values(sortby).reset_index(drop=True)

    ##############################################################
    # Formulas
    ##############################################################

    def col_name2col_letter(col_name):
        col_num = header.index(col_name) + 1
        return opx.utils.get_column_letter(col_num)

    col_unit = col_name2col_letter('Unit')
    col_billed = col_name2col_letter('Billed Impressions/UVs')
    col_base_rate = col_name2col_letter('Base Rate')
    col_prod_rate = col_name2col_letter('Baked-in Production')
    col_gross_rate = col_name2col_letter('Gross Rate')
    col_gross_rev_wo_prod = col_name2col_letter('Gross Site Revenue (Does Not Include Production Fee)')
    col_revshare = col_name2col_letter('RevShare')
    col_own_partner_dsp = col_name2col_letter('Own/Partner/DSP')
    col_delivered = col_name2col_letter('Delivered')

    def get_gross_rev_wo_prod(i):
        row = str(i + 2)
        output = '=IF(' + col_unit + row + '="CPM", '
        output += col_billed + row + '/1000*' + col_base_rate + row + ', '
        output += col_billed + row + '*' + col_base_rate + row + ')'
        return output

    def get_expense(row):
        row_index = str(row.name + 2)

        if (row['Net Site Expense'] == '') or pd.isnull(row['Net Site Expense']):
            return '=' + col_gross_rev_wo_prod + row_index + '*' + col_revshare + row_index

        non_standard_rate = str(row['Net Site Expense'])
        if row['Unit'] == 'CPM':
            return '=' + col_billed + row_index + '/1000*' + non_standard_rate
        else:
            return '=' + col_billed + row_index + '*' + non_standard_rate

    def get_prod_fee(i):
        row = str(i + 2)
        output = '=' + col_billed + row + '*' + col_prod_rate + row
        return output

    def get_gross_rev_w_prod(i):
        row = str(i + 2)
        output = '=IF(' + col_unit + row + '="CPM", '
        output += col_billed + row + '/1000*' + col_gross_rate + row + ', '
        output += col_billed + row + '*' + col_gross_rate + row + ')'
        return output

    def get_billed_gt_delivered(i):
        row = str(i + 2)
        output = '=IF(' + col_own_partner_dsp + row + '="DSP", '
        output += '"", ' + col_billed + row + '>' + col_delivered + row + ')'
        return output

    data['Gross Site Revenue (Does Not Include Production Fee)'] = [get_gross_rev_wo_prod(i) for i in range(len(data))]
    data['Net Site Expense'] = data.apply(lambda row: get_expense(row), axis=1)
    data['Production Fee'] = [get_prod_fee(i) for i in range(len(data))]
    data['Gross Revenue (Includes Production Fee)'] = [get_gross_rev_w_prod(i) for i in range(len(data))]
    data['Billed > Delivered'] = [get_billed_gt_delivered(i) for i in range(len(data))]

    ##############################################################
    # Output excel
    ##############################################################

    writer = pd.ExcelWriter(output_file_name)
    data.to_excel(writer, 'Billable', index=False)
    writer.save()

    return





